from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta
from pathlib import Path
import os
import shutil
import sqlite3
import subprocess
import sys
import threading
import tkinter as tk
import unicodedata
from tkinter import filedialog, messagebox, ttk

from config import (
    ACOMPANHAMENTOS_DIR,
    AJUDA_RESUMIDA_PATH,
    APP_NAME,
    APP_VERSION,
    BACKUP_DIR,
    EVIDENCIAS_DIR,
    GUIA_RAPIDO_PATH,
    MANUAL_EDITORIAL_PATH,
    MANUAL_USUARIO_PATH,
    SAIDAS_DIR,
    SOCIALIZACOES_DIR,
    slugify,
)
from database import Database
from models import (
    ACOMPANHAMENTO_CATEGORIAS,
    ACOMPANHAMENTO_STATUS_OPTIONS,
    AUTO_OBSERVATIONS,
    BUSCA_ATIVA_STATUS_OPTIONS,
    PRIORIDADE_OPTIONS,
    SOCIALIZACAO_STATUS_OPTIONS,
    STATUS_OPTIONS,
    TURMA_STATUS_OPTIONS,
    WEEKDAY_OPTIONS,
)
from relatorio import export_pdf, generate_docx, generate_financial_statement_docx


class DatePickerDialog(tk.Toplevel):
    WEEKDAY_LABELS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
    MONTH_LABELS = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]

    def __init__(
        self,
        master: tk.Misc,
        initial_date: datetime,
        on_select,
        preferred_weekday: int | None = None,
        holiday_resolver=None,
    ) -> None:
        super().__init__(master)
        self.title("Selecionar data")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self.on_select = on_select
        self.current_year = initial_date.year
        self.current_month = initial_date.month
        self.selected_day = initial_date.day
        self.preferred_weekday = preferred_weekday
        self.holiday_resolver = holiday_resolver or (lambda _year: {})

        self._build_layout()
        self._render_days()

        self.bind("<Escape>", lambda _event: self.destroy())
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build_layout(self) -> None:
        container = ttk.Frame(self, padding=12)
        container.pack(fill="both", expand=True)

        header = ttk.Frame(container)
        header.pack(fill="x", pady=(0, 8))

        ttk.Button(header, text="<", width=3, command=lambda: self._change_month(-1)).pack(
            side="left"
        )
        self.month_title_var = tk.StringVar()
        ttk.Label(
            header,
            textvariable=self.month_title_var,
            anchor="center",
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left", expand=True, padx=8)
        ttk.Button(header, text=">", width=3, command=lambda: self._change_month(1)).pack(
            side="right"
        )

        week_header = ttk.Frame(container)
        week_header.pack(fill="x")
        for index, label in enumerate(self.WEEKDAY_LABELS):
            ttk.Label(week_header, text=label, anchor="center", width=5).grid(
                row=0, column=index, padx=1, pady=(0, 2)
            )

        self.days_frame = ttk.Frame(container)
        self.days_frame.pack(fill="both", expand=True)

        footer = ttk.Frame(container)
        footer.pack(fill="x", pady=(10, 0))
        self.holiday_summary_var = tk.StringVar(value="")
        ttk.Label(
            footer,
            textvariable=self.holiday_summary_var,
            justify="left",
            wraplength=280,
        ).pack(side="left", padx=(0, 12))
        ttk.Button(footer, text="Hoje", command=self._select_today).pack(side="left")
        ttk.Button(footer, text="Cancelar", command=self.destroy).pack(side="right")

        legend = ttk.Frame(container)
        legend.pack(fill="x", pady=(8, 0))
        self._add_legend_item(legend, "#dbeafe", "#1d4ed8", "Dia padrão da turma")
        self._add_legend_item(legend, "#fecaca", "#7f1d1d", "Feriado nacional")
        self._add_legend_item(legend, "#c084fc", "#ffffff", "Dia da turma e feriado")

    def _add_legend_item(self, parent: ttk.Frame, bg: str, fg: str, text: str) -> None:
        item = ttk.Frame(parent)
        item.pack(side="left", padx=(0, 10))
        tk.Label(item, text="  ", bg=bg, fg=fg, relief="solid", borderwidth=1).pack(side="left")
        ttk.Label(item, text=text).pack(side="left", padx=(4, 0))

    def _change_month(self, delta: int) -> None:
        new_month = self.current_month + delta
        new_year = self.current_year
        if new_month == 0:
            new_month = 12
            new_year -= 1
        elif new_month == 13:
            new_month = 1
            new_year += 1
        self.current_month = new_month
        self.current_year = new_year
        self._render_days()

    def _render_days(self) -> None:
        for widget in self.days_frame.winfo_children():
            widget.destroy()

        self.month_title_var.set(
            f"{self.MONTH_LABELS[self.current_month - 1]} de {self.current_year}"
        )

        month_days = calendar.Calendar(firstweekday=0).monthdayscalendar(
            self.current_year, self.current_month
        )
        holiday_map = self.holiday_resolver(self.current_year)
        holiday_lines = []
        for row_index, week in enumerate(month_days):
            for column_index, day in enumerate(week):
                if day == 0:
                    ttk.Label(self.days_frame, text="", width=5).grid(
                        row=row_index, column=column_index, padx=1, pady=1
                    )
                    continue

                current_date = date(self.current_year, self.current_month, day)
                is_holiday = current_date in holiday_map
                is_preferred_weekday = self.preferred_weekday == column_index
                if is_holiday:
                    holiday_lines.append(f"{day:02d}/{self.current_month:02d} - {holiday_map[current_date]}")

                bg = "#f3f4f6"
                fg = "#111827"
                if is_holiday and is_preferred_weekday:
                    bg = "#a855f7"
                    fg = "#ffffff"
                elif is_holiday:
                    bg = "#fecaca"
                    fg = "#7f1d1d"
                elif is_preferred_weekday:
                    bg = "#dbeafe"
                    fg = "#1d4ed8"

                relief = "raised"
                borderwidth = 1
                if (
                    self.current_year == datetime.today().year
                    and self.current_month == datetime.today().month
                    and day == datetime.today().day
                ):
                    relief = "solid"
                    borderwidth = 2

                button = tk.Button(
                    self.days_frame,
                    text=str(day),
                    width=5,
                    bg=bg,
                    fg=fg,
                    activebackground=bg,
                    activeforeground=fg,
                    relief=relief,
                    borderwidth=borderwidth,
                    command=lambda selected_day=day: self._select_day(selected_day),
                )
                button.grid(row=row_index, column=column_index, padx=1, pady=1)

        self.holiday_summary_var.set(
            "Feriados no mês:\n" + "\n".join(holiday_lines) if holiday_lines else "Sem feriados nacionais neste mês."
        )

    def _select_day(self, day: int) -> None:
        selected_date = datetime(self.current_year, self.current_month, day)
        self.on_select(selected_date.strftime("%d/%m/%Y"))
        self.destroy()

    def _select_today(self) -> None:
        today = datetime.today()
        self.on_select(today.strftime("%d/%m/%Y"))
        self.destroy()


class MultiplicaApp(tk.Tk):
    TURMA_WEEKDAY_TO_INDEX = {
        "SEG": 0,
        "TER": 1,
        "QUA": 2,
        "QUI": 3,
        "SEX": 4,
        "SAB": 5,
        "DOM": 6,
    }
    TERMS_VERSION = "2026-07-06"
    TERMS_CONFIG_KEY = "terms_accepted_version"
    TERMS_ACCEPTED_AT_KEY = "terms_accepted_at"
    DISPLAY_LABELS = {
        "": "Todos",
        "abertas": "Abertas",
        "ativos": "Ativos",
        "todos": "Todos",
        "nao_iniciado": "Não iniciado",
        "em_contato": "Em contato",
        "aguardando_retorno": "Aguardando retorno",
        "localizado": "Localizado",
        "encerrado": "Encerrado",
        "curso_errado": "Curso errado",
        "turma_errada": "Turma errada",
        "consta_em_outra_turma": "Consta em outra turma",
        "nao_aparece_na_lista": "Não aparece na lista",
        "acesso": "Problema de acesso",
        "teams": "Problema no Teams",
        "inscricao": "Inscrição inconsistente",
        "busca_ativa": "Busca ativa",
        "duvida_pedagogica": "Dúvida pedagógica",
        "outro": "Outro",
        "aberto": "Aberto",
        "em_acompanhamento": "Em acompanhamento",
        "encaminhado_pec": "Encaminhado ao PEC",
        "resolvido": "Resolvido",
        "arquivado": "Arquivado",
        "baixa": "Baixa",
        "normal": "Normal",
        "alta": "Alta",
        "nao_enviada": "Não enviada",
        "enviada": "Enviada",
        "em_analise": "Em análise",
        "devolutiva_registrada": "Devolutiva registrada",
        "destaque": "Destaque",
        "com_anexos": "Com anexos",
        "sem_anexos": "Sem anexos",
        "com_movimentacoes": "Com movimentações",
        "sem_movimentacoes": "Sem movimentações",
    }

    def __init__(self, database: Database) -> None:
        super().__init__()
        self.database = database
        self.title(APP_NAME)
        self.geometry("1280x780")
        self.minsize(1080, 680)

        self.current_month_id: int | None = None
        self.selected_turma_id: int | None = None
        self.selected_encontro_id: int | None = None
        self.selected_evidence_id: int | None = None
        self.selected_cursista_id: int | None = None
        self.selected_acompanhamento_id: int | None = None
        self.selected_socializacao_id: int | None = None
        self.selected_acompanhamento_movimentacao_id: int | None = None
        self.selected_acompanhamento_anexo_id: int | None = None
        self.selected_socializacao_movimentacao_id: int | None = None
        self.selected_socializacao_anexo_id: int | None = None
        self.preview_photo = None
        self.preview_image_path: Path | None = None
        self.report_generation_in_progress = False

        self.month_label_to_id: dict[str, int] = {}
        self.turma_label_to_id: dict[str, int] = {}
        self.cursista_label_to_id: dict[str, int] = {}
        self.acompanhamento_encontro_label_to_id: dict[str, int] = {}
        self.socializacao_rows_by_item_id: dict[str, dict[str, object]] = {}
        self.acompanhamento_movimentacoes_rows: list[sqlite3.Row] = []
        self.acompanhamento_attachment_rows: dict[int, sqlite3.Row] = {}
        self.socializacao_movimentacoes_rows: list[sqlite3.Row] = []
        self.socializacao_attachment_rows: dict[int, sqlite3.Row] = {}
        self.auto_observacao_names = list(AUTO_OBSERVATIONS.keys())
        self._formatting_locks: set[str] = set()
        self._formatted_entries: dict[str, tuple[tk.Entry, object]] = {}
        self._tree_column_configs: dict[ttk.Treeview, dict[str, object]] = {}
        self._active_scroll_canvas: tk.Canvas | None = None

        self._build_variables()
        self._setup_auto_formatters()
        self._configure_styles()
        self._build_layout()
        self._load_initial_state()
        self.after(150, self.ensure_terms_acceptance)

    def _build_variables(self) -> None:
        today = datetime.today()

        self.current_month_label_var = tk.StringVar()

        self.nome_var = tk.StringVar()
        self.tema_var = tk.StringVar()
        self.email_var = tk.StringVar()
        self.diretoria_var = tk.StringVar()
        self.pec_var = tk.StringVar()
        self.valor_formacao_semanal_var = tk.StringVar()

        self.turma_codigo_var = tk.StringVar()
        self.turma_dia_var = tk.StringVar(value=WEEKDAY_OPTIONS[0])
        self.turma_horario_var = tk.StringVar()
        self.turma_componente_var = tk.StringVar()
        self.turma_situacao_var = tk.StringVar(value=TURMA_STATUS_OPTIONS[0])

        self.encontro_data_var = tk.StringVar()
        self.encontro_pauta_var = tk.StringVar()
        self.encontro_turma_var = tk.StringVar()
        self.encontro_inicio_var = tk.StringVar()
        self.encontro_fim_var = tk.StringVar()
        self.encontro_participantes_var = tk.StringVar()
        self.encontro_duracao_var = tk.StringVar()
        self.encontro_status_var = tk.StringVar(value=STATUS_OPTIONS[0])
        self.auto_observacao_var = tk.StringVar(value=self.auto_observacao_names[0])

        self.cursista_nome_var = tk.StringVar()
        self.cursista_turma_var = tk.StringVar()
        self.cursista_email_institucional_var = tk.StringVar()
        self.cursista_email_pessoal_var = tk.StringVar()
        self.cursista_telefone_var = tk.StringVar()
        self.cursista_status_busca_ativa_var = tk.StringVar(
            value=self.DISPLAY_LABELS[BUSCA_ATIVA_STATUS_OPTIONS[0]]
        )
        self.cursista_ativo_var = tk.BooleanVar(value=True)
        self.cursista_filter_nome_var = tk.StringVar()
        self.cursista_filter_turma_var = tk.StringVar()
        self.cursista_filter_ativo_var = tk.StringVar(value="ativos")
        self.cursista_filter_busca_var = tk.StringVar()

        self.acompanhamento_cursista_var = tk.StringVar()
        self.acompanhamento_turma_var = tk.StringVar()
        self.acompanhamento_encontro_var = tk.StringVar()
        self.acompanhamento_categoria_var = tk.StringVar(
            value=self._display_label(ACOMPANHAMENTO_CATEGORIAS[0])
        )
        self.acompanhamento_status_var = tk.StringVar(
            value=self._display_label(ACOMPANHAMENTO_STATUS_OPTIONS[0])
        )
        self.acompanhamento_prioridade_var = tk.StringVar(
            value=self._display_label(PRIORIDADE_OPTIONS[1])
        )
        self.acompanhamento_resumo_var = tk.StringVar()
        self.acompanhamento_data_abertura_var = tk.StringVar(value=today.strftime("%d/%m/%Y"))
        self.acompanhamento_data_resolucao_var = tk.StringVar()
        self.acompanhamento_encaminhar_pec_var = tk.BooleanVar(value=False)
        self.acompanhamento_filter_status_var = tk.StringVar(value=self.DISPLAY_LABELS["abertas"])
        self.acompanhamento_filter_turma_var = tk.StringVar()
        self.acompanhamento_filter_categoria_var = tk.StringVar()
        self.acompanhamento_filter_search_var = tk.StringVar()
        self.acompanhamento_filter_anexos_var = tk.StringVar()
        self.acompanhamento_filter_movimentacoes_var = tk.StringVar()
        self.acompanhamento_movimentacao_data_var = tk.StringVar(value=today.strftime("%d/%m/%Y"))

        self.socializacao_cursista_var = tk.StringVar()
        self.socializacao_turma_var = tk.StringVar()
        self.socializacao_mes_var = tk.StringVar(value=str(today.month))
        self.socializacao_ano_var = tk.StringVar(value=str(today.year))
        self.socializacao_status_var = tk.StringVar(
            value=self.DISPLAY_LABELS[SOCIALIZACAO_STATUS_OPTIONS[0]]
        )
        self.socializacao_data_envio_var = tk.StringVar()
        self.socializacao_necessita_apoio_var = tk.BooleanVar(value=False)
        self.socializacao_destaque_var = tk.BooleanVar(value=False)
        self.socializacao_filter_turma_var = tk.StringVar()
        self.socializacao_filter_status_var = tk.StringVar()
        self.socializacao_filter_search_var = tk.StringVar()
        self.socializacao_filter_anexos_var = tk.StringVar()
        self.socializacao_filter_movimentacoes_var = tk.StringVar()
        self.socializacao_movimentacao_data_var = tk.StringVar(value=today.strftime("%d/%m/%Y"))

        self.status_bar_var = tk.StringVar(
            value="Tudo permanece local: banco SQLite, imagens e relatórios."
        )

    def _setup_auto_formatters(self) -> None:
        self._formatted_entries = {}

    def _configure_styles(self) -> None:
        self.style = ttk.Style(self)
        try:
            if self.style.theme_use() != "clam":
                self.style.theme_use("clam")
        except tk.TclError:
            pass

        self.style.configure(
            "Multiplica.TNotebook",
            background="#dfe4ea",
            borderwidth=2,
            relief="solid",
            tabmargins=(4, 4, 4, 0),
        )
        self.style.configure(
            "Multiplica.TNotebook.Tab",
            padding=(12, 6, 12, 6),
            font=("Segoe UI", 9, "bold"),
            background="#e8edf3",
            foreground="#243447",
            borderwidth=1,
            relief="solid",
            lightcolor="#9aa9b8",
            darkcolor="#748394",
        )
        self.style.map(
            "Multiplica.TNotebook.Tab",
            background=[
                ("selected", "#ffffff"),
                ("active", "#f8fbff"),
            ],
            foreground=[
                ("selected", "#0f4c81"),
                ("active", "#12385f"),
            ],
            lightcolor=[
                ("selected", "#0f4c81"),
                ("active", "#4f83b8"),
            ],
            darkcolor=[
                ("selected", "#0f4c81"),
                ("active", "#4f83b8"),
            ],
            expand=[("selected", (1, 1, 1, 0))],
        )
        self.style.configure(
            "TButton",
            padding=(10, 5),
        )

    def _ttk_background(self, style_name: str = "TFrame") -> str:
        return self.style.lookup(style_name, "background") or "#f0f0f0"

    def _register_scroll_canvas(self, canvas: tk.Canvas, root_widget: tk.Misc) -> None:
        if not getattr(self, "_mousewheel_bound", False):
            self.bind_all("<MouseWheel>", self._handle_active_canvas_mousewheel, add="+")
            self.bind_all("<Button-4>", self._handle_active_canvas_mousewheel_linux, add="+")
            self.bind_all("<Button-5>", self._handle_active_canvas_mousewheel_linux, add="+")
            self._mousewheel_bound = True

        def _activate(_event=None) -> None:
            self._active_scroll_canvas = canvas

        def _deactivate(_event=None) -> None:
            if self._active_scroll_canvas is canvas:
                self._active_scroll_canvas = None

        widgets = [root_widget]
        widgets.extend(root_widget.winfo_children())
        while widgets:
            widget = widgets.pop()
            widget.bind("<Enter>", _activate, add="+")
            widget.bind("<Leave>", _deactivate, add="+")
            widgets.extend(widget.winfo_children())

    def _handle_active_canvas_mousewheel(self, event) -> None:
        canvas = self._active_scroll_canvas
        if canvas is None:
            return
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return
        if bool(event.state & 0x0001):
            canvas.xview_scroll(int(-delta / 120), "units")
            return
        canvas.yview_scroll(int(-delta / 120), "units")

    def _handle_active_canvas_mousewheel_linux(self, event) -> None:
        canvas = self._active_scroll_canvas
        if canvas is None:
            return
        if getattr(event, "num", None) == 4:
            canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            canvas.yview_scroll(1, "units")

    def _register_formatted_entry(self, entry: tk.Entry, formatter) -> None:
        widget_name = str(entry)
        self._formatted_entries[widget_name] = (entry, formatter)
        entry.bind("<KeyRelease>", self._handle_formatted_entry_keyrelease, add="+")

    def _handle_formatted_entry_keyrelease(self, event) -> None:
        entry = event.widget
        widget_name = str(entry)
        if widget_name not in self._formatted_entries or widget_name in self._formatting_locks:
            return

        formatter = self._formatted_entries[widget_name][1]
        current_value = entry.get()
        cursor_index = entry.index(tk.INSERT)
        digits_before_cursor = sum(1 for char in current_value[:cursor_index] if char.isdigit())
        formatted_value = formatter(current_value)
        if formatted_value == current_value:
            return

        new_cursor_index = self._find_cursor_position_from_digits(formatted_value, digits_before_cursor)

        self._formatting_locks.add(widget_name)
        try:
            entry.delete(0, tk.END)
            entry.insert(0, formatted_value)
            entry.icursor(new_cursor_index)
        finally:
            self._formatting_locks.discard(widget_name)

    def _find_cursor_position_from_digits(self, formatted_value: str, digits_before_cursor: int) -> int:
        if digits_before_cursor <= 0:
            return 0

        counted_digits = 0
        for index, char in enumerate(formatted_value):
            if char.isdigit():
                counted_digits += 1
                if counted_digits >= digits_before_cursor:
                    return index + 1
        return len(formatted_value)

    def _register_treeview_autofit(
        self,
        tree: ttk.Treeview,
        columns: tuple[str, ...],
        weights: dict[str, int],
        min_widths: dict[str, int],
    ) -> None:
        self._tree_column_configs[tree] = {
            "columns": columns,
            "weights": weights,
            "min_widths": min_widths,
        }
        tree.bind("<Configure>", lambda _event, current_tree=tree: self._autofit_treeview_columns(current_tree), add="+")
        self.after(50, lambda current_tree=tree: self._autofit_treeview_columns(current_tree))

    def _autofit_treeview_columns(self, tree: ttk.Treeview) -> None:
        config = self._tree_column_configs.get(tree)
        if not config:
            return

        columns = config["columns"]
        weights = config["weights"]
        min_widths = config["min_widths"]

        total_width = max(tree.winfo_width() - 6, 200)
        total_weight = sum(weights[column] for column in columns)
        assigned_width = 0

        for index, column in enumerate(columns):
            if index == len(columns) - 1:
                width = max(total_width - assigned_width, min_widths[column])
            else:
                width = max(int(total_width * weights[column] / total_weight), min_widths[column])
                assigned_width += width
            tree.column(column, width=width, stretch=True)

    def _format_time_live(self, value: str) -> str:
        digits = "".join(char for char in value if char.isdigit())[:4]
        if len(digits) <= 2:
            return digits
        return f"{digits[:2]}:{digits[2:]}"

    def _format_date_live(self, value: str) -> str:
        digits = "".join(char for char in value if char.isdigit())[:8]
        if len(digits) <= 2:
            return digits
        if len(digits) <= 4:
            return f"{digits[:2]}/{digits[2:]}"
        return f"{digits[:2]}/{digits[2:4]}/{digits[4:]}"

    def _parse_currency_to_cents(self, value: str) -> int:
        normalized = value.strip().replace("R$", "").replace(" ", "")
        if "," in normalized and "." in normalized:
            if normalized.rfind(",") > normalized.rfind("."):
                normalized = normalized.replace(".", "").replace(",", ".")
            else:
                normalized = normalized.replace(",", "")
        elif "," in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        if not normalized:
            return 0
        amount = float(normalized)
        return max(int(round(amount * 100)), 0)

    def _format_currency_from_cents(self, cents: int) -> str:
        value = max(cents, 0) / 100
        formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {formatted}"

    def _normalize_currency_field(self, _event=None) -> bool:
        raw_value = self.valor_formacao_semanal_var.get().strip()
        if not raw_value:
            return True
        try:
            cents = self._parse_currency_to_cents(raw_value)
        except ValueError:
            messagebox.showwarning(
                "Valor inválido",
                "Informe o valor da formação semanal em formato numérico, por exemplo: 150,00",
            )
            return False
        self.valor_formacao_semanal_var.set(self._format_currency_from_cents(cents))
        return True

    def _confirm_month_rate_update(self, month_id: int) -> bool | None:
        month = self.database.get_month(month_id)
        if not month:
            return True

        current_rate_cents = int(month["weekly_rate_cents"] or 0)
        informed_rate_cents = self._parse_currency_to_cents(self.valor_formacao_semanal_var.get())
        if current_rate_cents == informed_rate_cents:
            return True

        month_label = str(month["ref_label"] or f"{int(month['ref_month']):02d}/{int(month['ref_year'])}")
        saved_label = self._format_currency_from_cents(current_rate_cents)
        informed_label = self._format_currency_from_cents(informed_rate_cents)
        answer = messagebox.askyesnocancel(
            "Divergência no valor da formação semanal",
            (
                f"Mês de referência: {month_label}\n\n"
                f"Valor salvo neste mês: {saved_label}\n"
                f"Valor informado agora: {informed_label}\n\n"
                "Sim = atualizar o valor deste mês\n"
                "Não = manter o valor já salvo neste mês\n"
                "Cancelar = interromper a operação"
            ),
        )
        if answer is None:
            return None
        if answer is False:
            self.valor_formacao_semanal_var.set(saved_label)
            return False
        return True

    def _resolve_display_date(self, value: str) -> datetime:
        cleaned = value.strip()
        if not cleaned:
            if self.current_month_id:
                month_row = self.database.get_month(self.current_month_id)
                if month_row:
                    return datetime(int(month_row["ref_year"]), int(month_row["ref_month"]), 1)
            return datetime.today()

        for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(cleaned, date_format)
            except ValueError:
                continue
        return datetime.today()

    def open_encontro_calendar(self) -> None:
        initial_date = self._resolve_display_date(self.encontro_data_var.get())
        preferred_weekday = self._get_selected_turma_weekday_index()
        picker = DatePickerDialog(
            self,
            initial_date,
            self.encontro_data_var.set,
            preferred_weekday=preferred_weekday,
            holiday_resolver=self._get_brazil_national_holidays,
        )
        picker.focus_set()

    def _get_selected_turma_row(self) -> sqlite3.Row | None:
        turma_label = self.encontro_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label)
        if not turma_id:
            return None
        return self.database.get_turma(turma_id)

    def _get_selected_turma_weekday_index(self) -> int | None:
        turma = self._get_selected_turma_row()
        if not turma:
            return None
        return self.TURMA_WEEKDAY_TO_INDEX.get((turma["dia_semana"] or "").strip().upper())

    def _calculate_easter_sunday(self, year: int) -> date:
        a = year % 19
        b = year // 100
        c = year % 100
        d = b // 4
        e = b % 4
        f = (b + 8) // 25
        g = (b - f + 1) // 3
        h = (19 * a + b - d - g + 15) % 30
        i = c // 4
        k = c % 4
        l = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * l) // 451
        month = (h + l - 7 * m + 114) // 31
        day = ((h + l - 7 * m + 114) % 31) + 1
        return date(year, month, day)

    def _get_brazil_national_holidays(self, year: int) -> dict[date, str]:
        easter = self._calculate_easter_sunday(year)
        good_friday = easter - timedelta(days=2)
        return {
            date(year, 1, 1): "Confraternização Universal",
            good_friday: "Paixão de Cristo",
            date(year, 4, 21): "Tiradentes",
            date(year, 5, 1): "Dia do Trabalhador",
            date(year, 9, 7): "Independência do Brasil",
            date(year, 10, 12): "Nossa Senhora Aparecida",
            date(year, 11, 2): "Finados",
            date(year, 11, 15): "Proclamação da República",
            date(year, 11, 20): "Consciência Negra",
            date(year, 12, 25): "Natal",
        }

    def _build_layout(self) -> None:
        self._build_top_bar()

        self.notebook = ttk.Notebook(self, style="Multiplica.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.home_tab = ttk.Frame(self.notebook, padding=10)
        self.professor_tab = ttk.Frame(self.notebook, padding=10)
        self.turmas_tab = ttk.Frame(self.notebook, padding=10)
        self.encontros_tab = ttk.Frame(self.notebook, padding=10)
        self.checklist_tab = ttk.Frame(self.notebook, padding=10)
        self.acompanhamento_tab = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.home_tab, text="Início")
        self.notebook.add(self.professor_tab, text="Professor Multiplicador")
        self.notebook.add(self.turmas_tab, text="Turmas")
        self.notebook.add(self.encontros_tab, text="Encontros")
        self.notebook.add(self.checklist_tab, text="Checklist e relatório")
        self.notebook.add(self.acompanhamento_tab, text="Acompanhamento de Cursistas")

        self._build_home_tab()
        self._build_professor_tab()
        self._build_turmas_tab()
        self._build_encontros_tab()
        self._build_checklist_tab()
        self._build_acompanhamento_tab()

        status_bar = ttk.Label(
            self,
            textvariable=self.status_bar_var,
            anchor="w",
            padding=(8, 6),
        )
        status_bar.pack(fill="x", side="bottom")

    def _build_top_bar(self) -> None:
        frame = ttk.Frame(self, padding=(8, 8, 8, 4))
        frame.pack(fill="x")

        ttk.Label(frame, text="Meses com encontros").grid(row=0, column=0, sticky="w")
        self.month_combo = ttk.Combobox(
            frame,
            textvariable=self.current_month_label_var,
            state="readonly",
            width=24,
        )
        self.month_combo.grid(row=1, column=0, sticky="w", padx=(0, 12))
        self.month_combo.bind("<<ComboboxSelected>>", self._on_month_selected)

        ttk.Button(frame, text="Abrir saídas", command=self.open_output_folder).grid(
            row=1, column=1, padx=6
        )
        ttk.Button(frame, text="Abrir pasta das evidências", command=self.open_evidence_folder).grid(
            row=1, column=2, padx=6
        )
        ttk.Button(frame, text="Backup do banco", command=self.backup_database).grid(
            row=1, column=3, padx=6
        )
        ttk.Button(frame, text="Ajuda", command=self.show_help_dialog).grid(
            row=1, column=4, padx=6
        )
        ttk.Button(frame, text="Sobre o projeto", command=self.show_about_dialog).grid(
            row=1, column=5, padx=6
        )

        frame.columnconfigure(6, weight=1)

    def _build_home_tab(self) -> None:
        title = ttk.Label(
            self.home_tab,
            text="Multiplica Evidências",
            font=("Segoe UI", 16, "bold"),
        )
        title.pack(anchor="w")

        intro_text = tk.Text(
            self.home_tab,
            wrap="word",
            height=3,
            borderwidth=0,
            relief="flat",
            background=self._ttk_background(),
            font=("Segoe UI", 10),
        )
        intro_text.insert("end", "Aplicação local, ")
        intro_text.insert("end", "experimental e independente", "bold")
        intro_text.insert(
            "end",
            " para apoiar encontros pedagógicos, evidências e documentos mensais "
            "em diferentes contextos de uso.",
        )
        intro_text.tag_configure("bold", font=("Segoe UI", 10, "bold"))
        intro_text.config(state="disabled")
        intro_text.pack(anchor="w", fill="x", pady=(8, 12))
        ttk.Label(
            self.home_tab,
            text=f"Versão atual: {APP_VERSION}",
            justify="left",
        ).pack(anchor="w", pady=(0, 8))
        line_frame = ttk.LabelFrame(self.home_tab, text="Linha do projeto", padding=10)
        line_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(
            line_frame,
            text="Esta instalação corresponde à Linha 2 (edição avançada).",
            font=("Segoe UI", 10, "bold"),
            justify="left",
        ).pack(anchor="w")
        ttk.Label(
            line_frame,
            text=(
                "Linha estável 1.x: versão mais simples, indicada para uso mais direto no dia a dia.\n"
                "Linha avançada 2.x: versão com acompanhamento de cursistas, importação, anexos, "
                "movimentações e ajuda integrada.\n"
                "Os materiais de apoio desta instalação foram preparados para a Linha 2."
            ),
            justify="left",
        ).pack(anchor="w", pady=(6, 0))
        tk.Label(
            self.home_tab,
            text=(
                "AVISO IMPORTANTE: ESTE PROJETO É INDEPENDENTE E NÃO OFICIAL.\n"
                "Pode ser utilizado experimentalmente em encontros pedagógicos, inclusive em contextos como o\n"
                "Programa Multiplica, sem representar homologação ou vínculo institucional automático."
            ),
            font=("Segoe UI", 10, "bold"),
            fg="#8b1e1e",
            bg=self._ttk_background(),
            justify="left",
        ).pack(anchor="w", pady=(0, 8))
        ttk.Label(
            self.home_tab,
            text=(
                "Aviso: no primeiro uso, o sistema solicita o aceite dos termos de uso "
                "antes da utilização."
            ),
            justify="left",
        ).pack(anchor="w", pady=(0, 12))

        quick = ttk.LabelFrame(self.home_tab, text="Fluxo sugerido", padding=10)
        quick.pack(fill="x", pady=(0, 10))
        ttk.Label(
            quick,
            text=(
                "1. Salve os dados do professor.\n"
                "2. Cadastre as turmas.\n"
                "3. Registre cada encontro e anexe os prints.\n"
                "4. O mês é criado automaticamente pela data do encontro.\n"
                "5. Selecione um mês existente para revisar o checklist e gerar o relatório."
            ),
            justify="left",
        ).pack(anchor="w")

        actions = ttk.LabelFrame(self.home_tab, text="Ações rápidas", padding=10)
        actions.pack(fill="x")
        buttons = [
            ("Cadastrar professor", 1),
            ("Cadastrar turma", 2),
            ("Registrar encontro", 3),
            ("Checklist mensal", 4),
            ("Sobre o projeto", None),
        ]
        for column, (label, tab_index) in enumerate(buttons):
            ttk.Button(
                actions,
                text=label,
                command=(
                    self.show_about_dialog
                    if tab_index is None
                    else lambda index=tab_index: self.notebook.select(index)
                ),
            ).grid(row=0, column=column, padx=(0, 8), pady=2, sticky="w")

        help_frame = ttk.LabelFrame(self.home_tab, text="Ajuda e documentação", padding=10)
        help_frame.pack(fill="x", pady=(10, 0))
        ttk.Label(
            help_frame,
            text=(
                "Se preferir, consulte o guia rápido, o manual completo ou a versão editorial "
                "do manual diretamente por esta tela."
            ),
            justify="left",
        ).pack(anchor="w", pady=(0, 8))

        help_buttons = ttk.Frame(help_frame)
        help_buttons.pack(fill="x")
        ttk.Button(help_buttons, text="Abrir ajuda", command=self.show_help_dialog).pack(side="left")
        ttk.Button(help_buttons, text="Guia rápido", command=self.open_quick_guide).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(help_buttons, text="Manual completo", command=self.open_user_manual).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(help_buttons, text="Pasta docs", command=self.open_docs_folder).pack(
            side="left", padx=(8, 0)
        )

    def _build_professor_tab(self) -> None:
        frame = ttk.LabelFrame(self.professor_tab, text="Dados do professor multiplicador", padding=10)
        frame.pack(fill="x", anchor="n")

        fields = [
            ("Nome", self.nome_var, 34),
            ("Tema", self.tema_var, 34),
            ("E-mail institucional", self.email_var, 42),
            ("Diretoria / URE", self.diretoria_var, 28),
            ("PEC responsável", self.pec_var, 24),
            ("Valor por formação semanal (R$)", self.valor_formacao_semanal_var, 14),
        ]
        for row, (label, variable, width) in enumerate(fields):
            ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", pady=4)
            entry = ttk.Entry(frame, textvariable=variable, width=width)
            entry.grid(row=row, column=1, sticky="w", pady=4, padx=(12, 0))
            if variable is self.valor_formacao_semanal_var:
                entry.bind("<FocusOut>", self._normalize_currency_field, add="+")

        frame.columnconfigure(1, weight=1)

        button_row = ttk.Frame(self.professor_tab, padding=(0, 10, 0, 0))
        button_row.pack(fill="x")
        ttk.Button(button_row, text="Salvar dados do professor", command=self.save_current_month).pack(
            side="left"
        )
        ttk.Button(
            button_row,
            text="Salvar como padrão",
            command=self.save_defaults,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(
            button_row,
            text="Usar padrão salvo",
            command=self.apply_defaults_to_form,
        ).pack(side="left", padx=(10, 0))

    def _build_turmas_tab(self) -> None:
        top = ttk.Frame(self.turmas_tab)
        top.pack(fill="both", expand=True)

        form = ttk.LabelFrame(top, text="Cadastro de turma", padding=10)
        form.pack(fill="x", anchor="n")

        fields = [
            ("Código da turma", self.turma_codigo_var, 24),
            ("Dia da semana", self.turma_dia_var, 12),
            ("Horário", self.turma_horario_var, 10),
            ("Tema / componente", self.turma_componente_var, 30),
            ("Situação", self.turma_situacao_var, 12),
        ]

        ttk.Label(form, text=fields[0][0]).grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=fields[0][1], width=fields[0][2]).grid(
            row=0, column=1, sticky="w", pady=4, padx=(12, 16)
        )

        ttk.Label(form, text=fields[1][0]).grid(row=0, column=2, sticky="w", pady=4)
        ttk.Combobox(
            form,
            values=WEEKDAY_OPTIONS,
            textvariable=fields[1][1],
            state="readonly",
            width=10,
        ).grid(row=0, column=3, sticky="w", pady=4, padx=(12, 16))

        ttk.Label(form, text=fields[2][0]).grid(row=0, column=4, sticky="w", pady=4)
        turma_horario_entry = ttk.Entry(form, textvariable=fields[2][1], width=fields[2][2])
        turma_horario_entry.grid(row=0, column=5, sticky="w", pady=4, padx=(12, 16))
        self._register_formatted_entry(turma_horario_entry, self._format_time_live)

        ttk.Label(form, text=fields[3][0]).grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=fields[3][1], width=fields[3][2]).grid(
            row=1, column=1, columnspan=3, sticky="w", pady=4, padx=(12, 16)
        )

        ttk.Label(form, text=fields[4][0]).grid(row=1, column=4, sticky="w", pady=4)
        ttk.Combobox(
            form,
            values=TURMA_STATUS_OPTIONS,
            textvariable=fields[4][1],
            state="readonly",
            width=10,
        ).grid(row=1, column=5, sticky="w", pady=4, padx=(12, 0))

        buttons = ttk.Frame(form)
        buttons.grid(row=2, column=0, columnspan=6, sticky="w", pady=(8, 0))
        ttk.Button(buttons, text="Salvar turma", command=self.save_turma).pack(side="left")
        ttk.Button(buttons, text="Nova turma", command=self.clear_turma_form).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(
            buttons,
            text="Carregar selecionada",
            command=self.load_selected_turma,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(
            buttons,
            text="Alternar situação",
            command=self.toggle_selected_turma_status,
        ).pack(side="left", padx=(10, 0))

        tree_frame = ttk.LabelFrame(top, text="Turmas cadastradas", padding=8)
        tree_frame.pack(fill="both", expand=True, pady=(10, 0))

        columns = ("codigo", "dia", "horario", "componente", "situacao")
        self.turma_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        headers = {
            "codigo": "Código",
            "dia": "Dia",
            "horario": "Horário",
            "componente": "Componente",
            "situacao": "Situação",
        }
        min_widths = {
            "codigo": 150,
            "dia": 70,
            "horario": 80,
            "componente": 150,
            "situacao": 90,
        }
        weights = {
            "codigo": 34,
            "dia": 10,
            "horario": 12,
            "componente": 28,
            "situacao": 16,
        }
        for column in columns:
            self.turma_tree.heading(column, text=headers[column])
            self.turma_tree.column(column, width=min_widths[column], anchor="w", stretch=True)
        self.turma_tree.pack(side="left", fill="both", expand=True)
        self.turma_tree.bind("<Double-1>", lambda _event: self.load_selected_turma())
        self._register_treeview_autofit(self.turma_tree, columns, weights, min_widths)

        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.turma_tree.yview)
        tree_scroll.pack(side="right", fill="y")
        self.turma_tree.configure(yscrollcommand=tree_scroll.set)

    def _build_encontros_tab(self) -> None:
        paned = ttk.Panedwindow(self.encontros_tab, orient="horizontal")
        paned.pack(fill="both", expand=True)

        left = ttk.Frame(paned, padding=(0, 0, 8, 0))
        right = ttk.Frame(paned)
        paned.add(left, weight=2)
        paned.add(right, weight=2)

        form = ttk.LabelFrame(left, text="Registro de encontro", padding=10)
        form.pack(fill="x", anchor="n")

        ttk.Label(form, text="Turma").grid(row=0, column=0, sticky="w", pady=4)
        self.encontro_turma_combo = ttk.Combobox(
            form,
            textvariable=self.encontro_turma_var,
            state="readonly",
            width=28,
        )
        self.encontro_turma_combo.grid(
            row=0, column=1, columnspan=4, sticky="w", pady=4, padx=(12, 0)
        )
        self.encontro_turma_combo.bind("<<ComboboxSelected>>", self._on_encontro_turma_selected)

        ttk.Label(form, text="Data").grid(
            row=1, column=0, sticky="w", pady=4
        )
        encontro_data_entry = ttk.Entry(form, textvariable=self.encontro_data_var, width=16)
        encontro_data_entry.grid(row=1, column=1, sticky="w", padx=(12, 8), pady=4)
        self._register_formatted_entry(encontro_data_entry, self._format_date_live)
        ttk.Button(form, text="📅", width=3, command=self.open_encontro_calendar).grid(
            row=1, column=2, sticky="w", pady=4, padx=(0, 16)
        )
        ttk.Label(form, text="Número da pauta").grid(row=1, column=3, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.encontro_pauta_var, width=8).grid(
            row=1, column=4, sticky="w", padx=(12, 0), pady=4
        )

        ttk.Label(form, text="Início").grid(row=2, column=0, sticky="w", pady=4)
        encontro_inicio_entry = ttk.Entry(form, textvariable=self.encontro_inicio_var, width=10)
        encontro_inicio_entry.grid(row=2, column=1, sticky="w", padx=(12, 16), pady=4)
        self._register_formatted_entry(encontro_inicio_entry, self._format_time_live)
        ttk.Label(form, text="Término").grid(row=2, column=3, sticky="w", pady=4)
        encontro_fim_entry = ttk.Entry(form, textvariable=self.encontro_fim_var, width=10)
        encontro_fim_entry.grid(row=2, column=4, sticky="w", padx=(12, 0), pady=4)
        self._register_formatted_entry(encontro_fim_entry, self._format_time_live)

        ttk.Label(form, text="Participantes").grid(row=3, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.encontro_participantes_var, width=10).grid(
            row=3, column=1, sticky="w", padx=(12, 16), pady=4
        )
        ttk.Label(form, text="Duração").grid(row=3, column=3, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.encontro_duracao_var, width=10).grid(
            row=3, column=4, sticky="w", padx=(12, 0), pady=4
        )

        ttk.Label(form, text="Situação").grid(row=4, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form,
            values=STATUS_OPTIONS,
            textvariable=self.encontro_status_var,
            state="readonly",
            width=16,
        ).grid(row=4, column=1, sticky="w", padx=(12, 16), pady=4)

        ttk.Label(form, text="Texto automático").grid(row=4, column=3, sticky="w", pady=4)
        ttk.Combobox(
            form,
            values=self.auto_observacao_names,
            textvariable=self.auto_observacao_var,
            state="readonly",
            width=18,
        ).grid(row=4, column=4, sticky="w", padx=(12, 0), pady=4)

        ttk.Button(form, text="Aplicar texto", command=self.apply_auto_observation).grid(
            row=5, column=4, sticky="e", pady=(6, 8)
        )

        ttk.Label(form, text="Observação").grid(row=5, column=0, sticky="nw", pady=(10, 4))
        self.observacao_text = tk.Text(form, width=62, height=5, wrap="word")
        self.observacao_text.grid(row=6, column=0, columnspan=5, sticky="ew", pady=(0, 8))

        action_row = ttk.Frame(form)
        action_row.grid(row=7, column=0, columnspan=5, sticky="w")
        ttk.Button(action_row, text="Salvar encontro", command=self.save_encontro).pack(side="left")
        ttk.Button(action_row, text="Inserir encontro", command=self.clear_encontro_form).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(
            action_row,
            text="Carregar selecionado",
            command=self.load_selected_encontro,
        ).pack(side="left", padx=(10, 0))

        form.columnconfigure(1, weight=1)
        form.columnconfigure(4, weight=1)

        evidence_frame = ttk.LabelFrame(left, text="Evidências do encontro", padding=10)
        evidence_frame.pack(fill="both", expand=True, pady=(10, 0))

        evidence_list_frame = ttk.Frame(evidence_frame)
        evidence_list_frame.pack(fill="both", expand=True)

        self.evidence_listbox = tk.Listbox(evidence_list_frame, height=9)
        self.evidence_listbox.pack(side="left", fill="both", expand=True)
        self.evidence_listbox.bind("<<ListboxSelect>>", self._on_evidence_selected)

        evidence_scroll_y = ttk.Scrollbar(
            evidence_list_frame, orient="vertical", command=self.evidence_listbox.yview
        )
        evidence_scroll_y.pack(side="right", fill="y")
        self.evidence_listbox.configure(yscrollcommand=evidence_scroll_y.set)

        evidence_buttons = ttk.Frame(evidence_frame)
        evidence_buttons.pack(fill="x", pady=(10, 0))
        ttk.Button(evidence_buttons, text="Colar print", command=self.paste_image_from_clipboard).pack(
            side="left"
        )
        ttk.Button(evidence_buttons, text="Adicionar imagem", command=self.add_image_file).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(
            evidence_buttons,
            text="Visualizar imagem",
            command=self.open_selected_evidence,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(evidence_buttons, text="Remover imagem", command=self.remove_selected_evidence).pack(
            side="left", padx=(10, 0)
        )

        right_paned = ttk.Panedwindow(right, orient="vertical")
        right_paned.pack(fill="both", expand=True)

        encontro_list_frame = ttk.LabelFrame(right_paned, text="Encontros do mês", padding=8)
        preview_frame = ttk.LabelFrame(right_paned, text="Pré-visualização da evidência", padding=8)
        right_paned.add(encontro_list_frame, weight=1)
        right_paned.add(preview_frame, weight=1)

        columns = ("data", "pauta", "turma", "status", "participantes", "imagens")
        self.encontro_tree = ttk.Treeview(
            encontro_list_frame,
            columns=columns,
            show="headings",
            height=15,
        )
        headers = {
            "data": "Data",
            "pauta": "Pauta",
            "turma": "Turma",
            "status": "Situação",
            "participantes": "Participantes",
            "imagens": "Imagens",
        }
        min_widths = {
            "data": 80,
            "pauta": 70,
            "turma": 115,
            "status": 100,
            "participantes": 88,
            "imagens": 68,
        }
        weights = {
            "data": 12,
            "pauta": 10,
            "turma": 28,
            "status": 20,
            "participantes": 16,
            "imagens": 14,
        }
        for column in columns:
            self.encontro_tree.heading(column, text=headers[column])
            self.encontro_tree.column(column, width=min_widths[column], anchor="w", stretch=True)
        self.encontro_tree.pack(side="left", fill="both", expand=True)
        self.encontro_tree.bind("<Double-1>", lambda _event: self.load_selected_encontro())
        self._register_treeview_autofit(self.encontro_tree, columns, weights, min_widths)

        encontro_scroll = ttk.Scrollbar(
            encontro_list_frame, orient="vertical", command=self.encontro_tree.yview
        )
        encontro_scroll.pack(side="right", fill="y")
        encontro_scroll_x = ttk.Scrollbar(
            encontro_list_frame, orient="horizontal", command=self.encontro_tree.xview
        )
        encontro_scroll_x.pack(side="bottom", fill="x")
        self.encontro_tree.configure(
            yscrollcommand=encontro_scroll.set,
            xscrollcommand=encontro_scroll_x.set,
        )

        self.preview_title_var = tk.StringVar(value="Selecione uma evidência à esquerda.")
        ttk.Label(preview_frame, textvariable=self.preview_title_var).pack(anchor="w", pady=(0, 8))

        self.preview_canvas = tk.Canvas(preview_frame, background="#f3f4f6", highlightthickness=0)
        self.preview_canvas.pack(fill="both", expand=True)
        self.preview_canvas.bind("<Configure>", lambda _event: self._render_preview_image())

    def _build_checklist_tab(self) -> None:
        summary = ttk.LabelFrame(self.checklist_tab, text="Conferência mensal", padding=10)
        summary.pack(fill="x")

        self.checklist_labels: dict[str, ttk.Label] = {}
        rows = [
            ("total_registrados", "Total registrados"),
            ("total_realizados", "Total realizados"),
            ("valor_formacao_semanal", "Valor formação semanal"),
            ("valor_total_mensal", "Valor total mensal"),
            ("turmas", "Turmas"),
            ("pautas", "Pautas"),
            ("sem_imagem", "Sem imagem"),
            ("sem_observacao", "Sem observação"),
            ("sem_horario", "Sem horário"),
            ("sem_participantes", "Sem participantes"),
        ]
        for row, (key, label_text) in enumerate(rows):
            ttk.Label(summary, text=label_text).grid(
                row=row, column=0, sticky="w", pady=2
            )
            label = ttk.Label(summary, text="-")
            label.grid(row=row, column=1, sticky="w", padx=(12, 0), pady=2)
            self.checklist_labels[key] = label

        detail_frame = ttk.LabelFrame(
            self.checklist_tab,
            text="Pendências e detalhes",
            padding=10,
        )
        detail_frame.pack(fill="both", expand=True, pady=(10, 0))

        self.checklist_listbox = tk.Listbox(detail_frame, height=14)
        self.checklist_listbox.pack(fill="both", expand=True)

        action_row = ttk.Frame(self.checklist_tab, padding=(0, 10, 0, 0))
        action_row.pack(fill="x")
        self.refresh_checklist_button = ttk.Button(
            action_row, text="Atualizar checklist", command=self.refresh_checklist
        )
        self.refresh_checklist_button.pack(
            side="left"
        )
        self.generate_docx_button = ttk.Button(
            action_row, text="Gerar relatório DOCX", command=self.generate_docx_report
        )
        self.generate_docx_button.pack(side="left", padx=(10, 0))
        self.generate_pdf_button = ttk.Button(
            action_row, text="Gerar relatório PDF", command=self.generate_pdf_report
        )
        self.generate_pdf_button.pack(side="left", padx=(10, 0))
        self.generate_financial_button = ttk.Button(
            action_row,
            text="Gerar extrato financeiro",
            command=self.generate_financial_report,
        )
        self.generate_financial_button.pack(side="left", padx=(10, 0))
        self.open_evidence_button = ttk.Button(
            action_row,
            text="Abrir pasta das evidências",
            command=self.open_evidence_folder,
        )
        self.open_evidence_button.pack(side="left", padx=(10, 0))
        self.open_output_button = ttk.Button(
            action_row, text="Abrir pasta de saída", command=self.open_output_folder
        )
        self.open_output_button.pack(side="left", padx=(10, 0))

        self.report_progress_var = tk.StringVar(value="")
        self.report_progress = ttk.Progressbar(
            self.checklist_tab, mode="indeterminate", length=260
        )
        self.report_progress.pack(anchor="w", pady=(10, 0))
        self.report_progress.pack_forget()
        self.report_progress_label = ttk.Label(
            self.checklist_tab, textvariable=self.report_progress_var
        )
        self.report_progress_label.pack(anchor="w", pady=(6, 0))

    def _build_acompanhamento_tab(self) -> None:
        dashboard = ttk.LabelFrame(
            self.acompanhamento_tab,
            text="Resumo do acompanhamento",
            padding=10,
        )
        dashboard.pack(fill="x")
        self.cursistas_dashboard_labels: dict[str, ttk.Label] = {}
        dashboard_rows = [
            ("cursistas_ativos", "Cursistas ativos"),
            ("busca_ativa_pendente", "Busca ativa pendente"),
            ("acompanhamentos_abertos", "Ocorrências abertas"),
            ("socializacoes_nao_enviadas", "Socializações não enviadas"),
            ("socializacoes_destaque", "Potenciais destaques"),
        ]
        for index, (key, label_text) in enumerate(dashboard_rows):
            column = index * 2
            ttk.Label(dashboard, text=label_text).grid(
                row=0,
                column=column,
                sticky="w",
                padx=(0 if index == 0 else 18, 8),
                pady=2,
            )
            label = ttk.Label(dashboard, text="-")
            label.grid(row=0, column=column + 1, sticky="w", pady=2)
            self.cursistas_dashboard_labels[key] = label
        for column in range(len(dashboard_rows) * 2):
            dashboard.columnconfigure(column, weight=1 if column % 2 == 0 else 0)

        self.acompanhamento_notebook = ttk.Notebook(
            self.acompanhamento_tab,
            style="Multiplica.TNotebook",
        )
        self.acompanhamento_notebook.pack(fill="both", expand=True, pady=(10, 0))

        self.cursistas_subtab = ttk.Frame(self.acompanhamento_notebook, padding=10)
        self.ocorrencias_subtab = ttk.Frame(self.acompanhamento_notebook, padding=10)
        self.socializacoes_subtab = ttk.Frame(self.acompanhamento_notebook, padding=10)

        self.acompanhamento_notebook.add(self.cursistas_subtab, text="Cursistas")
        self.acompanhamento_notebook.add(self.ocorrencias_subtab, text="Ocorrências")
        self.acompanhamento_notebook.add(self.socializacoes_subtab, text="Socializações")

        self._build_cursistas_subtab()
        self._build_ocorrencias_subtab()
        self._build_socializacoes_subtab()

    def _build_cursistas_subtab(self) -> None:
        filter_frame = ttk.LabelFrame(self.cursistas_subtab, text="Filtros", padding=10)
        filter_frame.pack(fill="x")
        ttk.Label(filter_frame, text="Nome").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(filter_frame, textvariable=self.cursista_filter_nome_var, width=28).grid(
            row=0, column=1, sticky="w", padx=(8, 16), pady=4
        )
        ttk.Label(filter_frame, text="Turma").grid(row=0, column=2, sticky="w", pady=4)
        self.cursista_filter_turma_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.cursista_filter_turma_var,
            state="readonly",
            width=18,
        )
        self.cursista_filter_turma_combo.grid(row=0, column=3, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Busca ativa").grid(row=0, column=4, sticky="w", pady=4)
        self.cursista_filter_busca_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.cursista_filter_busca_var,
            state="readonly",
            width=20,
            values=self._display_options([""] + BUSCA_ATIVA_STATUS_OPTIONS),
        )
        self.cursista_filter_busca_combo.grid(row=0, column=5, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Situação").grid(row=0, column=6, sticky="w", pady=4)
        self.cursista_filter_ativo_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.cursista_filter_ativo_var,
            state="readonly",
            width=12,
            values=["ativos", "todos"],
        )
        self.cursista_filter_ativo_combo.grid(row=0, column=7, sticky="w", padx=(8, 16), pady=4)
        ttk.Button(filter_frame, text="Atualizar", command=self.refresh_cursista_tree).grid(
            row=0, column=8, sticky="w", pady=4
        )
        ttk.Button(
            filter_frame,
            text="Analisar planilha",
            command=self.preview_cursista_import_file,
        ).grid(row=0, column=9, sticky="w", pady=4)

        content = ttk.Panedwindow(self.cursistas_subtab, orient="horizontal")
        content.pack(fill="both", expand=True, pady=(10, 0))

        left = ttk.Frame(content)
        right = ttk.Frame(content)
        content.add(left, weight=2)
        content.add(right, weight=3)

        form = ttk.LabelFrame(left, text="Cadastro de cursista", padding=10)
        form.pack(fill="both", expand=True)

        ttk.Label(form, text="Nome").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.cursista_nome_var, width=34).grid(
            row=0, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=4
        )
        ttk.Label(form, text="Turma").grid(row=1, column=0, sticky="w", pady=4)
        self.cursista_turma_combo = ttk.Combobox(
            form,
            textvariable=self.cursista_turma_var,
            state="readonly",
            width=24,
        )
        self.cursista_turma_combo.grid(row=1, column=1, sticky="w", padx=(10, 16), pady=4)
        ttk.Label(form, text="Busca ativa").grid(row=1, column=2, sticky="w", pady=4)
        ttk.Combobox(
            form,
            textvariable=self.cursista_status_busca_ativa_var,
            values=self._display_options(BUSCA_ATIVA_STATUS_OPTIONS),
            state="readonly",
            width=22,
        ).grid(row=1, column=3, sticky="w", padx=(10, 0), pady=4)
        ttk.Label(form, text="E-mail institucional").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.cursista_email_institucional_var, width=34).grid(
            row=2, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=4
        )
        ttk.Label(form, text="E-mail pessoal").grid(row=3, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.cursista_email_pessoal_var, width=34).grid(
            row=3, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=4
        )
        ttk.Label(form, text="Telefone / WhatsApp").grid(row=4, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.cursista_telefone_var, width=24).grid(
            row=4, column=1, sticky="w", padx=(10, 16), pady=4
        )
        ttk.Checkbutton(form, text="Cursista ativo", variable=self.cursista_ativo_var).grid(
            row=4, column=2, columnspan=2, sticky="w", pady=4
        )
        ttk.Label(form, text="Observação geral").grid(row=6, column=0, sticky="nw", pady=(4, 4))
        self.cursista_observacao_text = tk.Text(form, width=44, height=5, wrap="word")
        self.cursista_observacao_text.grid(
            row=7, column=0, columnspan=4, sticky="nsew", pady=(0, 6)
        )

        button_row = ttk.Frame(form)
        button_row.grid(row=5, column=0, columnspan=4, sticky="w", pady=(10, 6))
        ttk.Button(button_row, text="Salvar cursista", command=self.save_cursista).pack(side="left")
        ttk.Button(button_row, text="Novo cursista", command=self.clear_cursista_form).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(
            button_row,
            text="Carregar selecionado",
            command=self.load_selected_cursista,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(
            button_row,
            text="Inativar",
            command=self.deactivate_selected_cursista,
        ).pack(side="left", padx=(10, 0))

        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)
        form.rowconfigure(7, weight=1)

        tree_frame = ttk.LabelFrame(right, text="Cursistas cadastrados", padding=8)
        tree_frame.pack(fill="both", expand=True)
        columns = ("nome", "turma", "busca", "email", "telefone", "ativo")
        self.cursista_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=16)
        headers = {
            "nome": "Nome",
            "turma": "Turma",
            "busca": "Busca ativa",
            "email": "E-mail",
            "telefone": "Telefone",
            "ativo": "Ativo",
        }
        min_widths = {
            "nome": 180,
            "turma": 90,
            "busca": 130,
            "email": 180,
            "telefone": 110,
            "ativo": 60,
        }
        weights = {
            "nome": 28,
            "turma": 12,
            "busca": 18,
            "email": 24,
            "telefone": 12,
            "ativo": 6,
        }
        for column in columns:
            self.cursista_tree.heading(column, text=headers[column])
            self.cursista_tree.column(column, width=min_widths[column], anchor="w", stretch=True)
        self.cursista_tree.pack(side="left", fill="both", expand=True)
        self.cursista_tree.bind("<Double-1>", lambda _event: self.load_selected_cursista())
        self._register_treeview_autofit(self.cursista_tree, columns, weights, min_widths)
        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.cursista_tree.yview)
        scroll.pack(side="right", fill="y")
        self.cursista_tree.configure(yscrollcommand=scroll.set)

    def _build_ocorrencias_subtab(self) -> None:
        filter_frame = ttk.LabelFrame(self.ocorrencias_subtab, text="Filtros", padding=10)
        filter_frame.pack(fill="x")
        ttk.Label(filter_frame, text="Status").grid(row=0, column=0, sticky="w", pady=4)
        self.acompanhamento_filter_status_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.acompanhamento_filter_status_var,
            state="readonly",
            width=18,
            values=self._display_options(["abertas", ""] + ACOMPANHAMENTO_STATUS_OPTIONS),
        )
        self.acompanhamento_filter_status_combo.grid(
            row=0, column=1, sticky="w", padx=(8, 16), pady=4
        )
        ttk.Label(filter_frame, text="Turma").grid(row=0, column=2, sticky="w", pady=4)
        self.acompanhamento_filter_turma_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.acompanhamento_filter_turma_var,
            state="readonly",
            width=18,
        )
        self.acompanhamento_filter_turma_combo.grid(
            row=0, column=3, sticky="w", padx=(8, 16), pady=4
        )
        self.acompanhamento_filter_turma_combo.bind(
            "<<ComboboxSelected>>",
            lambda _event: self._refresh_cursista_combo_options(),
        )
        ttk.Label(filter_frame, text="Categoria").grid(row=0, column=4, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.acompanhamento_filter_categoria_var,
            values=self._display_options([""] + ACOMPANHAMENTO_CATEGORIAS),
            state="readonly",
            width=20,
        ).grid(row=0, column=5, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Busca").grid(row=0, column=6, sticky="w", pady=4)
        ttk.Entry(filter_frame, textvariable=self.acompanhamento_filter_search_var, width=24).grid(
            row=0, column=7, sticky="w", padx=(8, 16), pady=4
        )
        ttk.Label(filter_frame, text="Anexos").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.acompanhamento_filter_anexos_var,
            values=self._display_options(["", "com_anexos", "sem_anexos"]),
            state="readonly",
            width=18,
        ).grid(row=1, column=1, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Movimentações").grid(row=1, column=2, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.acompanhamento_filter_movimentacoes_var,
            values=self._display_options(["", "com_movimentacoes", "sem_movimentacoes"]),
            state="readonly",
            width=18,
        ).grid(row=1, column=3, sticky="w", padx=(8, 16), pady=4)
        ttk.Button(
            filter_frame,
            text="Atualizar",
            command=self.refresh_acompanhamento_tree,
        ).grid(row=1, column=8, sticky="w", pady=4)

        content = ttk.Panedwindow(self.ocorrencias_subtab, orient="horizontal")
        content.pack(fill="both", expand=True, pady=(10, 0))
        left = ttk.Frame(content)
        right = ttk.Frame(content)
        content.add(left, weight=2)
        content.add(right, weight=3)

        left_canvas = tk.Canvas(
            left,
            background=self._ttk_background(),
            highlightthickness=0,
        )
        left_scrollbar = ttk.Scrollbar(left, orient="vertical", command=left_canvas.yview)
        left_scrollbar.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)
        left_canvas.configure(yscrollcommand=left_scrollbar.set)

        left_content = ttk.Frame(left_canvas)
        left_window = left_canvas.create_window((0, 0), window=left_content, anchor="nw")

        def _update_acompanhamento_form_scroll(_event=None) -> None:
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))

        def _resize_acompanhamento_form_width(event) -> None:
            left_canvas.itemconfigure(left_window, width=event.width)

        left_content.bind("<Configure>", _update_acompanhamento_form_scroll)
        left_canvas.bind("<Configure>", _resize_acompanhamento_form_width)
        self._register_scroll_canvas(left_canvas, left_content)

        form = ttk.LabelFrame(left_content, text="Ocorrência do cursista", padding=10)
        form.pack(fill="both", expand=True)
        ttk.Label(form, text="Cursista").grid(row=0, column=0, sticky="w", pady=4)
        self.acompanhamento_cursista_combo = ttk.Combobox(
            form,
            textvariable=self.acompanhamento_cursista_var,
            state="readonly",
            width=28,
        )
        self.acompanhamento_cursista_combo.grid(
            row=0, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=4
        )
        self.acompanhamento_cursista_combo.bind(
            "<<ComboboxSelected>>",
            self._on_acompanhamento_cursista_selected,
        )
        ttk.Label(form, text="Turma").grid(row=1, column=0, sticky="w", pady=4)
        self.acompanhamento_turma_combo = ttk.Combobox(
            form,
            textvariable=self.acompanhamento_turma_var,
            state="readonly",
            width=18,
        )
        self.acompanhamento_turma_combo.grid(row=1, column=1, sticky="w", padx=(10, 16), pady=4)
        self.acompanhamento_turma_combo.bind(
            "<<ComboboxSelected>>",
            self._on_acompanhamento_turma_selected,
        )
        ttk.Label(form, text="Encontro").grid(row=1, column=2, sticky="w", pady=4)
        self.acompanhamento_encontro_combo = ttk.Combobox(
            form,
            textvariable=self.acompanhamento_encontro_var,
            state="readonly",
            width=22,
        )
        self.acompanhamento_encontro_combo.grid(row=1, column=3, sticky="w", padx=(10, 0), pady=4)
        ttk.Label(form, text="Categoria").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form,
            textvariable=self.acompanhamento_categoria_var,
            values=self._display_options(ACOMPANHAMENTO_CATEGORIAS),
            state="readonly",
            width=22,
        ).grid(row=2, column=1, sticky="w", padx=(10, 16), pady=4)
        ttk.Label(form, text="Status").grid(row=2, column=2, sticky="w", pady=4)
        ttk.Combobox(
            form,
            textvariable=self.acompanhamento_status_var,
            values=self._display_options(ACOMPANHAMENTO_STATUS_OPTIONS),
            state="readonly",
            width=20,
        ).grid(row=2, column=3, sticky="w", padx=(10, 0), pady=4)
        ttk.Label(form, text="Prioridade").grid(row=3, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form,
            textvariable=self.acompanhamento_prioridade_var,
            values=self._display_options(PRIORIDADE_OPTIONS),
            state="readonly",
            width=12,
        ).grid(row=3, column=1, sticky="w", padx=(10, 16), pady=4)
        ttk.Label(form, text="Abertura").grid(row=3, column=2, sticky="w", pady=4)
        data_abertura_entry = ttk.Entry(
            form,
            textvariable=self.acompanhamento_data_abertura_var,
            width=14,
        )
        data_abertura_entry.grid(row=3, column=3, sticky="w", padx=(10, 0), pady=4)
        self._register_formatted_entry(data_abertura_entry, self._format_date_live)
        ttk.Label(form, text="Resumo").grid(row=4, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.acompanhamento_resumo_var, width=40).grid(
            row=4, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=4
        )
        ttk.Label(form, text="Resolução").grid(row=5, column=0, sticky="w", pady=4)
        data_resolucao_entry = ttk.Entry(
            form,
            textvariable=self.acompanhamento_data_resolucao_var,
            width=14,
        )
        data_resolucao_entry.grid(row=5, column=1, sticky="w", padx=(10, 16), pady=4)
        self._register_formatted_entry(data_resolucao_entry, self._format_date_live)
        ttk.Checkbutton(
            form,
            text="Encaminhar ao PEC",
            variable=self.acompanhamento_encaminhar_pec_var,
        ).grid(row=5, column=2, columnspan=2, sticky="w", pady=4)
        ttk.Label(form, text="Descrição").grid(row=6, column=0, sticky="nw", pady=(10, 4))
        self.acompanhamento_descricao_text = tk.Text(form, width=44, height=8, wrap="word")
        self.acompanhamento_descricao_text.grid(
            row=7, column=0, columnspan=4, sticky="ew", pady=(0, 8)
        )
        button_row = ttk.Frame(form)
        button_row.grid(row=8, column=0, columnspan=4, sticky="w")
        ttk.Button(
            button_row,
            text="Salvar ocorrência",
            command=self.save_acompanhamento,
        ).pack(side="left")
        ttk.Button(
            button_row,
            text="Nova ocorrência",
            command=self.clear_acompanhamento_form,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(
            button_row,
            text="Carregar selecionada",
            command=self.load_selected_acompanhamento,
        ).pack(side="left", padx=(10, 0))

        movimentacao_frame = ttk.LabelFrame(form, text="Movimentações", padding=8)
        movimentacao_frame.grid(row=9, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        ttk.Label(movimentacao_frame, text="Data").grid(row=0, column=0, sticky="w", pady=4)
        movimento_data_entry = ttk.Entry(
            movimentacao_frame,
            textvariable=self.acompanhamento_movimentacao_data_var,
            width=14,
        )
        movimento_data_entry.grid(row=0, column=1, sticky="w", padx=(8, 12), pady=4)
        self._register_formatted_entry(movimento_data_entry, self._format_date_live)
        ttk.Button(
            movimentacao_frame,
            text="Adicionar movimentação",
            command=self.add_acompanhamento_movimentacao,
        ).grid(row=0, column=2, sticky="w", pady=4)
        self.acompanhamento_movimentacao_text = tk.Text(
            movimentacao_frame,
            width=42,
            height=3,
            wrap="word",
        )
        self.acompanhamento_movimentacao_text.grid(
            row=1,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(0, 6),
        )
        self.acompanhamento_movimentacoes_listbox = tk.Listbox(movimentacao_frame, height=4)
        self.acompanhamento_movimentacoes_listbox.grid(
            row=2,
            column=0,
            columnspan=3,
            sticky="ew",
        )
        ttk.Button(
            movimentacao_frame,
            text="Remover movimentação",
            command=self.remove_selected_acompanhamento_movimentacao,
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(6, 0))

        anexos_frame = ttk.LabelFrame(form, text="Anexos da ocorrência", padding=8)
        anexos_frame.grid(row=10, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        self.acompanhamento_anexos_listbox = tk.Listbox(anexos_frame, height=4)
        self.acompanhamento_anexos_listbox.grid(row=0, column=0, columnspan=3, sticky="ew")
        ttk.Button(
            anexos_frame,
            text="Adicionar arquivo",
            command=self.add_acompanhamento_attachment,
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Button(
            anexos_frame,
            text="Abrir arquivo",
            command=self.open_selected_acompanhamento_attachment,
        ).grid(row=1, column=1, sticky="w", padx=(10, 0), pady=(6, 0))
        ttk.Button(
            anexos_frame,
            text="Remover arquivo",
            command=self.remove_selected_acompanhamento_attachment,
        ).grid(row=1, column=2, sticky="w", padx=(10, 0), pady=(6, 0))
        resumo_frame = ttk.LabelFrame(form, text="Resumo da ocorrência", padding=8)
        resumo_frame.grid(row=11, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        self.acompanhamento_summary_labels: dict[str, ttk.Label] = {}
        for row_index, (key, label_text) in enumerate(
            [
                ("status", "Status"),
                ("periodo", "Abertura / resolução"),
                ("movimentacoes", "Movimentações"),
                ("anexos", "Anexos"),
            ]
        ):
            ttk.Label(resumo_frame, text=label_text).grid(row=row_index, column=0, sticky="w", pady=2)
            label = ttk.Label(resumo_frame, text="-")
            label.grid(row=row_index, column=1, sticky="w", padx=(10, 0), pady=2)
            self.acompanhamento_summary_labels[key] = label
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        tree_frame = ttk.LabelFrame(right, text="Ocorrências registradas", padding=8)
        tree_frame.pack(fill="both", expand=True)
        columns = ("abertura", "cursista", "turma", "categoria", "status", "prioridade")
        self.acompanhamento_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            height=16,
        )
        headers = {
            "abertura": "Abertura",
            "cursista": "Cursista",
            "turma": "Turma",
            "categoria": "Categoria",
            "status": "Status",
            "prioridade": "Prioridade",
        }
        min_widths = {
            "abertura": 90,
            "cursista": 180,
            "turma": 80,
            "categoria": 130,
            "status": 120,
            "prioridade": 90,
        }
        weights = {
            "abertura": 12,
            "cursista": 28,
            "turma": 10,
            "categoria": 22,
            "status": 18,
            "prioridade": 10,
        }
        for column in columns:
            self.acompanhamento_tree.heading(column, text=headers[column])
            self.acompanhamento_tree.column(column, width=min_widths[column], anchor="w", stretch=True)
        self.acompanhamento_tree.pack(side="left", fill="both", expand=True)
        self.acompanhamento_tree.tag_configure("pending", background="#fff7d6")
        self.acompanhamento_tree.tag_configure("resolved", background="#e7f7ec")
        self.acompanhamento_tree.tag_configure("high_priority", background="#fde2e2")
        self.acompanhamento_tree.bind(
            "<Double-1>",
            lambda _event: self.load_selected_acompanhamento(),
        )
        self._register_treeview_autofit(self.acompanhamento_tree, columns, weights, min_widths)
        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.acompanhamento_tree.yview)
        scroll.pack(side="right", fill="y")
        self.acompanhamento_tree.configure(yscrollcommand=scroll.set)

    def _build_socializacoes_subtab(self) -> None:
        filter_frame = ttk.LabelFrame(self.socializacoes_subtab, text="Filtros do mês", padding=10)
        filter_frame.pack(fill="x")
        ttk.Label(filter_frame, text="Mês").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.socializacao_mes_var,
            values=[str(item) for item in range(1, 13)],
            state="readonly",
            width=6,
        ).grid(row=0, column=1, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Ano").grid(row=0, column=2, sticky="w", pady=4)
        ttk.Entry(filter_frame, textvariable=self.socializacao_ano_var, width=8).grid(
            row=0, column=3, sticky="w", padx=(8, 16), pady=4
        )
        ttk.Label(filter_frame, text="Turma").grid(row=0, column=4, sticky="w", pady=4)
        self.socializacao_filter_turma_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.socializacao_filter_turma_var,
            state="readonly",
            width=18,
        )
        self.socializacao_filter_turma_combo.grid(
            row=0, column=5, sticky="w", padx=(8, 16), pady=4
        )
        self.socializacao_filter_turma_combo.bind(
            "<<ComboboxSelected>>",
            lambda _event: self._refresh_cursista_combo_options(),
        )
        ttk.Label(filter_frame, text="Status").grid(row=0, column=6, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.socializacao_filter_status_var,
            values=self._display_options([""] + SOCIALIZACAO_STATUS_OPTIONS),
            state="readonly",
            width=20,
        ).grid(row=0, column=7, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Busca").grid(row=0, column=8, sticky="w", pady=4)
        ttk.Entry(filter_frame, textvariable=self.socializacao_filter_search_var, width=20).grid(
            row=0, column=9, sticky="w", padx=(8, 16), pady=4
        )
        ttk.Label(filter_frame, text="Anexos").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.socializacao_filter_anexos_var,
            values=self._display_options(["", "com_anexos", "sem_anexos"]),
            state="readonly",
            width=18,
        ).grid(row=1, column=1, sticky="w", padx=(8, 16), pady=4)
        ttk.Label(filter_frame, text="Movimentações").grid(row=1, column=2, sticky="w", pady=4)
        ttk.Combobox(
            filter_frame,
            textvariable=self.socializacao_filter_movimentacoes_var,
            values=self._display_options(["", "com_movimentacoes", "sem_movimentacoes"]),
            state="readonly",
            width=18,
        ).grid(row=1, column=3, sticky="w", padx=(8, 16), pady=4)
        ttk.Button(
            filter_frame,
            text="Atualizar",
            command=self.refresh_socializacao_tree,
        ).grid(row=1, column=10, sticky="w", pady=4)

        content = ttk.Panedwindow(self.socializacoes_subtab, orient="horizontal")
        content.pack(fill="both", expand=True, pady=(10, 0))
        left = ttk.Frame(content)
        right = ttk.Frame(content)
        content.add(left, weight=3)
        content.add(right, weight=2)

        tree_frame = ttk.LabelFrame(left, text="Socializações do mês", padding=8)
        tree_frame.pack(fill="both", expand=True)
        columns = ("cursista", "turma", "status", "envio", "apoio", "destaque")
        self.socializacao_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=11)
        headers = {
            "cursista": "Cursista",
            "turma": "Turma",
            "status": "Status",
            "envio": "Envio",
            "apoio": "Apoio",
            "destaque": "Destaque",
        }
        min_widths = {
            "cursista": 180,
            "turma": 90,
            "status": 140,
            "envio": 90,
            "apoio": 70,
            "destaque": 80,
        }
        weights = {
            "cursista": 32,
            "turma": 12,
            "status": 24,
            "envio": 14,
            "apoio": 9,
            "destaque": 9,
        }
        for column in columns:
            self.socializacao_tree.heading(column, text=headers[column])
            self.socializacao_tree.column(column, width=min_widths[column], anchor="w", stretch=True)
        self.socializacao_tree.pack(side="left", fill="both", expand=True)
        self.socializacao_tree.tag_configure("pending", background="#fff7d6")
        self.socializacao_tree.tag_configure("support", background="#fdecc8")
        self.socializacao_tree.tag_configure("highlight", background="#e9f6e9")
        self.socializacao_tree.bind("<Double-1>", lambda _event: self.load_selected_socializacao())
        self._register_treeview_autofit(self.socializacao_tree, columns, weights, min_widths)
        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.socializacao_tree.yview)
        scroll.pack(side="right", fill="y")
        self.socializacao_tree.configure(yscrollcommand=scroll.set)

        right_canvas = tk.Canvas(
            right,
            background=self._ttk_background(),
            highlightthickness=0,
        )
        right_scrollbar = ttk.Scrollbar(right, orient="vertical", command=right_canvas.yview)
        right_scrollbar.pack(side="right", fill="y")
        right_canvas.pack(side="left", fill="both", expand=True)
        right_canvas.configure(yscrollcommand=right_scrollbar.set)

        right_content = ttk.Frame(right_canvas)
        right_window = right_canvas.create_window((0, 0), window=right_content, anchor="nw")

        def _update_socializacao_form_scroll(_event=None) -> None:
            right_canvas.configure(scrollregion=right_canvas.bbox("all"))

        def _resize_socializacao_form_width(event) -> None:
            right_canvas.itemconfigure(right_window, width=event.width)

        right_content.bind("<Configure>", _update_socializacao_form_scroll)
        right_canvas.bind("<Configure>", _resize_socializacao_form_width)
        self._register_scroll_canvas(right_canvas, right_content)

        form = ttk.LabelFrame(right_content, text="Registro da socialização", padding=10)
        form.pack(fill="both", expand=True)
        ttk.Label(form, text="Cursista").grid(row=0, column=0, sticky="w", pady=4)
        self.socializacao_cursista_combo = ttk.Combobox(
            form,
            textvariable=self.socializacao_cursista_var,
            state="readonly",
            width=28,
        )
        self.socializacao_cursista_combo.grid(
            row=0, column=1, sticky="ew", padx=(10, 0), pady=4
        )
        self.socializacao_cursista_combo.bind(
            "<<ComboboxSelected>>",
            self._on_socializacao_cursista_selected,
        )
        ttk.Label(form, text="Turma").grid(row=1, column=0, sticky="w", pady=4)
        self.socializacao_turma_combo = ttk.Combobox(
            form,
            textvariable=self.socializacao_turma_var,
            state="readonly",
            width=22,
        )
        self.socializacao_turma_combo.grid(row=1, column=1, sticky="w", padx=(10, 0), pady=4)
        self.socializacao_turma_combo.bind(
            "<<ComboboxSelected>>",
            self._on_socializacao_turma_selected,
        )
        ttk.Label(form, text="Status").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form,
            textvariable=self.socializacao_status_var,
            values=self._display_options(SOCIALIZACAO_STATUS_OPTIONS),
            state="readonly",
            width=22,
        ).grid(row=2, column=1, sticky="w", padx=(10, 0), pady=4)
        ttk.Label(form, text="Data de envio").grid(row=3, column=0, sticky="w", pady=4)
        data_envio_entry = ttk.Entry(form, textvariable=self.socializacao_data_envio_var, width=14)
        data_envio_entry.grid(row=3, column=1, sticky="w", padx=(10, 0), pady=4)
        self._register_formatted_entry(data_envio_entry, self._format_date_live)
        ttk.Checkbutton(
            form,
            text="Necessita apoio",
            variable=self.socializacao_necessita_apoio_var,
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=4)
        ttk.Checkbutton(
            form,
            text="Potencial destaque",
            variable=self.socializacao_destaque_var,
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=4)
        ttk.Label(form, text="Observação pedagógica").grid(row=7, column=0, sticky="nw", pady=(4, 4))
        self.socializacao_observacao_text = tk.Text(form, width=34, height=4, wrap="word")
        self.socializacao_observacao_text.grid(row=8, column=0, columnspan=2, sticky="nsew", pady=(0, 6))
        button_row = ttk.Frame(form)
        button_row.grid(row=6, column=0, columnspan=2, sticky="w", pady=(8, 6))
        ttk.Button(button_row, text="Salvar socialização", command=self.save_socializacao).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Button(button_row, text="Nova ficha", command=self.clear_socializacao_form).grid(
            row=0, column=1, sticky="w", padx=(10, 0)
        )
        ttk.Button(
            button_row,
            text="Carregar selecionada",
            command=self.load_selected_socializacao,
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Button(
            button_row,
            text="Marcar enviada hoje",
            command=self.mark_socializacao_sent_today,
        ).grid(row=1, column=1, sticky="w", padx=(10, 0), pady=(6, 0))

        movimentacao_frame = ttk.LabelFrame(form, text="Movimentações", padding=8)
        movimentacao_frame.grid(row=9, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Label(movimentacao_frame, text="Data").grid(row=0, column=0, sticky="w", pady=4)
        socializacao_movimento_data_entry = ttk.Entry(
            movimentacao_frame,
            textvariable=self.socializacao_movimentacao_data_var,
            width=14,
        )
        socializacao_movimento_data_entry.grid(row=0, column=1, sticky="w", padx=(8, 12), pady=4)
        self._register_formatted_entry(socializacao_movimento_data_entry, self._format_date_live)
        ttk.Button(
            movimentacao_frame,
            text="Adicionar movimentação",
            command=self.add_socializacao_movimentacao,
        ).grid(row=0, column=2, sticky="w", pady=4)
        self.socializacao_movimentacao_text = tk.Text(
            movimentacao_frame,
            width=34,
            height=2,
            wrap="word",
        )
        self.socializacao_movimentacao_text.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 6))
        self.socializacao_movimentacoes_listbox = tk.Listbox(movimentacao_frame, height=2)
        self.socializacao_movimentacoes_listbox.grid(row=2, column=0, columnspan=3, sticky="ew")
        ttk.Button(
            movimentacao_frame,
            text="Remover movimentação",
            command=self.remove_selected_socializacao_movimentacao,
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(6, 0))

        anexos_frame = ttk.LabelFrame(form, text="Anexos da socialização", padding=8)
        anexos_frame.grid(row=10, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        self.socializacao_anexos_listbox = tk.Listbox(anexos_frame, height=2)
        self.socializacao_anexos_listbox.grid(row=0, column=0, columnspan=3, sticky="ew")
        ttk.Button(
            anexos_frame,
            text="Adicionar arquivo",
            command=self.add_socializacao_attachment,
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Button(
            anexos_frame,
            text="Abrir arquivo",
            command=self.open_selected_socializacao_attachment,
        ).grid(row=1, column=1, sticky="w", padx=(10, 0), pady=(6, 0))
        ttk.Button(
            anexos_frame,
            text="Remover arquivo",
            command=self.remove_selected_socializacao_attachment,
        ).grid(row=1, column=2, sticky="w", padx=(10, 0), pady=(6, 0))
        resumo_frame = ttk.LabelFrame(form, text="Resumo da socialização", padding=8)
        resumo_frame.grid(row=11, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        self.socializacao_summary_labels: dict[str, ttk.Label] = {}
        for row_index, (key, label_text) in enumerate(
            [
                ("status", "Status"),
                ("envio", "Envio"),
                ("movimentacoes", "Movimentações"),
                ("anexos", "Anexos"),
            ]
        ):
            ttk.Label(resumo_frame, text=label_text).grid(row=row_index, column=0, sticky="w", pady=2)
            label = ttk.Label(resumo_frame, text="-")
            label.grid(row=row_index, column=1, sticky="w", padx=(10, 0), pady=2)
            self.socializacao_summary_labels[key] = label
        form.columnconfigure(1, weight=1)
        form.rowconfigure(8, weight=1)

    def _load_initial_state(self) -> None:
        self.apply_defaults_to_form()
        self.refresh_month_selector()
        self.refresh_turma_tree()
        self.refresh_encontro_tree()
        self.refresh_checklist()
        self._refresh_cursista_combo_options()
        self.refresh_cursista_tree()
        self.refresh_acompanhamento_tree()
        self.refresh_socializacao_tree()

        if self.month_label_to_id:
            first_label = next(iter(self.month_label_to_id))
            self.current_month_label_var.set(first_label)
            self.load_month(self.month_label_to_id[first_label])
        else:
            self.current_month_id = None
            self.current_month_label_var.set("")
            self.status_bar_var.set(
                "Nenhum mês com encontros ainda. O primeiro mês será criado ao salvar um encontro."
            )
        self._refresh_cursistas_dashboard()

    def refresh_month_selector(self, select_month_id: int | None = None) -> None:
        rows = self.database.list_months()
        self.month_label_to_id = {row["ref_label"]: int(row["id"]) for row in rows}
        labels = list(self.month_label_to_id.keys())
        self.month_combo["values"] = labels

        if select_month_id is not None:
            for label, month_id in self.month_label_to_id.items():
                if month_id == select_month_id:
                    self.current_month_label_var.set(label)
                    return

        if labels and not self.current_month_label_var.get():
            self.current_month_label_var.set(labels[0])

    def _on_month_selected(self, _event=None) -> None:
        label = self.current_month_label_var.get()
        month_id = self.month_label_to_id.get(label)
        if month_id:
            self.load_month(month_id)

    def load_month(self, month_id: int) -> None:
        row = self.database.get_month(month_id)
        if not row:
            return

        self.current_month_id = month_id
        self.nome_var.set(row["teacher_name"])
        self.tema_var.set(row["theme"])
        self.email_var.set(row["institutional_email"])
        self.diretoria_var.set(row["diretoria_ure"])
        self.pec_var.set(row["pec_responsavel"])
        self.valor_formacao_semanal_var.set(self._format_currency_from_cents(int(row["weekly_rate_cents"] or 0)))
        self.status_bar_var.set(
            f"Mês ativo: {row['ref_label']} | Dados mantidos apenas localmente."
        )

        self.clear_encontro_form()
        self.refresh_turma_tree()
        self.refresh_encontro_tree()
        self.refresh_checklist()
        self._refresh_cursistas_dashboard()
        self.refresh_acompanhamento_tree()
        self.refresh_socializacao_tree()

    def _format_cursista_label(self, row: sqlite3.Row | dict[str, object]) -> str:
        turma_codigo = ""
        if isinstance(row, sqlite3.Row):
            turma_codigo = str(row["turma_codigo"] or "").strip()
            nome = str(row["nome"] or "").strip()
            identifier = int(row["id"])
        else:
            turma_codigo = str(row.get("turma_codigo", "") or "").strip()
            nome = str(row.get("nome", "") or "").strip()
            identifier = int(row["id"])
        suffix = turma_codigo or "sem turma"
        return f"{nome} | {suffix} | #{identifier}"

    def _display_label(self, value: str) -> str:
        return self.DISPLAY_LABELS.get(value, value.replace("_", " ").strip().title())

    def _display_options(self, values: list[str]) -> list[str]:
        return [self._display_label(value) for value in values]

    def _internal_value_from_label(self, label: str) -> str:
        normalized = label.strip()
        for key, value in self.DISPLAY_LABELS.items():
            if value == normalized:
                return key
        return normalized

    def _normalize_import_header(self, value: object) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(char for char in text if not unicodedata.combining(char))
        return " ".join(text.replace("_", " ").replace("-", " ").split())

    def _display_import_preview_status(self, value: str) -> str:
        mapping = {
            "novo": "Novo",
            "atualizar": "Atualizar",
            "conflito": "Conflito",
            "ignorado": "Ignorado",
        }
        return mapping.get(value, value.replace("_", " ").strip().title())

    def _display_import_preview_reason(self, value: str) -> str:
        normalized = str(value or "").strip()
        mapping = {
            "": "",
            "linha_vazia": "Linha vazia",
            "nome_nao_identificado": "Nome não identificado",
            "turma_nao_identificada": "Turma não identificada",
            "turma_nao_cadastrada": "Turma não cadastrada",
            "mais_de_um_cursista_mesmo_nome_turma": "Mais de um cursista com mesmo nome e turma",
            "email_ja_cadastrado": "E-mail já cadastrado em outro cursista",
            "cursista_nao_encontrado_na_aplicacao": "Cursista não encontrado na aplicação",
        }
        if normalized in mapping:
            return mapping[normalized]
        if normalized.startswith("duplicada_na_planilha_linha_"):
            line_number = normalized.rsplit("_", 1)[-1]
            return f"Duplicada na planilha (linha {line_number})"
        return normalized.replace("_", " ").strip().capitalize()

    def _format_encontro_label(self, row: sqlite3.Row | dict[str, object]) -> str:
        data = self._format_date_for_display(str(row["data_encontro"]), full_year=True)
        return f"{data} | Pauta {row['pauta_numero']} | {row['turma_codigo']}"

    def _mask_email(self, value: str) -> str:
        text = value.strip()
        if not text or "@" not in text:
            return text
        local, domain = text.split("@", 1)
        if len(local) <= 2:
            masked_local = local[:1] + "*"
        else:
            masked_local = local[:2] + "*" * max(len(local) - 2, 1)
        return f"{masked_local}@{domain}"

    def _mask_phone(self, value: str) -> str:
        digits = "".join(char for char in value if char.isdigit())
        if len(digits) < 4:
            return value.strip()
        return f"***{digits[-4:]}"

    def _parse_optional_date(self, raw_value: str) -> str:
        return self._parse_date(raw_value) if raw_value.strip() else ""

    def _get_turma_values_for_combos(self) -> list[str]:
        active_labels = list(self.encontro_turma_combo["values"]) if hasattr(self, "encontro_turma_combo") else []
        available = active_labels or list(self.turma_label_to_id.keys())
        return [""] + list(available)

    def _refresh_cursista_combo_options(self) -> None:
        self.cursista_label_to_id = {}
        all_rows = self.database.list_cursistas(include_inactive=False)
        for row in all_rows:
            label = self._format_cursista_label(row)
            self.cursista_label_to_id[label] = int(row["id"])

        combo_configs: list[tuple[ttk.Combobox, str]] = []
        if hasattr(self, "acompanhamento_cursista_combo"):
            turma_label = (
                self.acompanhamento_turma_var.get().strip()
                or self.acompanhamento_filter_turma_var.get().strip()
            )
            combo_configs.append((self.acompanhamento_cursista_combo, turma_label))
        if hasattr(self, "socializacao_cursista_combo"):
            turma_label = (
                self.socializacao_turma_var.get().strip()
                or self.socializacao_filter_turma_var.get().strip()
            )
            combo_configs.append((self.socializacao_cursista_combo, turma_label))

        for combo, turma_label in combo_configs:
            turma_id = self.turma_label_to_id.get(turma_label) if turma_label else None
            rows = self.database.list_cursistas(include_inactive=False, turma_id=turma_id)
            labels = [self._format_cursista_label(row) for row in rows]
            combo["values"] = labels
            current_value = combo.get().strip()
            if current_value and current_value not in labels:
                combo.set("")

    def _refresh_cursista_module_turma_options(self) -> None:
        values = self._get_turma_values_for_combos()
        combo_names = [
            "cursista_turma_combo",
            "cursista_filter_turma_combo",
            "acompanhamento_turma_combo",
            "acompanhamento_filter_turma_combo",
            "socializacao_turma_combo",
            "socializacao_filter_turma_combo",
        ]
        for name in combo_names:
            combo = getattr(self, name, None)
            if combo is not None:
                combo["values"] = values
                current_value = combo.get().strip()
                if current_value and current_value not in values:
                    combo.set("")

    def _populate_acompanhamento_encontros_for_turma(
        self,
        turma_id: int | None = None,
        month_id: int | None = None,
    ) -> None:
        self.acompanhamento_encontro_label_to_id = {}
        values = [""]
        if turma_id:
            rows = self.database.list_encontros_for_turma(turma_id, month_id or self.current_month_id)
            for row in rows:
                label = self._format_encontro_label(row)
                values.append(label)
                self.acompanhamento_encontro_label_to_id[label] = int(row["id"])
        self.acompanhamento_encontro_combo["values"] = values
        if self.acompanhamento_encontro_var.get().strip() not in values:
            self.acompanhamento_encontro_var.set("")

    def _refresh_cursistas_dashboard(self) -> None:
        try:
            month = int(self.socializacao_mes_var.get() or datetime.today().month)
            year = int(self.socializacao_ano_var.get() or datetime.today().year)
        except ValueError:
            month = datetime.today().month
            year = datetime.today().year
        counts = self.database.get_cursistas_dashboard_counts(month, year)
        for key, label in self.cursistas_dashboard_labels.items():
            label.configure(text=str(counts.get(key, 0)))

    def _build_acompanhamento_attachment_directory(self, acompanhamento_id: int) -> Path:
        target = ACOMPANHAMENTOS_DIR / f"acomp_{acompanhamento_id:05d}"
        target.mkdir(parents=True, exist_ok=True)
        return target

    def _build_socializacao_attachment_directory(self, socializacao_id: int) -> Path:
        target = SOCIALIZACOES_DIR / f"socializacao_{socializacao_id:05d}"
        target.mkdir(parents=True, exist_ok=True)
        return target

    def _next_attachment_path(self, folder: Path, source_name: str) -> Path:
        source_path = Path(source_name)
        stem = slugify(source_path.stem or "arquivo")
        extension = source_path.suffix or ".bin"
        counter = 1
        while True:
            candidate = folder / f"{stem}_{counter:02d}{extension}"
            if not candidate.exists():
                return candidate
            counter += 1

    def _delete_file_if_exists(self, file_path: str | Path) -> None:
        path = Path(file_path)
        if path.exists():
            path.unlink()

    def refresh_acompanhamento_movimentacoes_list(self) -> None:
        self.acompanhamento_movimentacoes_rows = []
        self.acompanhamento_movimentacoes_listbox.delete(0, "end")
        if not self.selected_acompanhamento_id:
            self._update_acompanhamento_summary()
            return
        rows = self.database.list_acompanhamento_movimentacoes(self.selected_acompanhamento_id)
        self.acompanhamento_movimentacoes_rows = rows
        for row in rows:
            data_label = self._format_date_for_display(str(row["data_movimentacao"]), full_year=True)
            resumo = str(row["descricao"] or "").strip().replace("\n", " ")
            label = f"{data_label} - {resumo[:80]}"
            self.acompanhamento_movimentacoes_listbox.insert("end", label)
        self._update_acompanhamento_summary()

    def refresh_acompanhamento_attachments_list(self) -> None:
        self.acompanhamento_attachment_rows = {}
        self.acompanhamento_anexos_listbox.delete(0, "end")
        if not self.selected_acompanhamento_id:
            self._update_acompanhamento_summary()
            return
        rows = self.database.list_acompanhamento_anexos(self.selected_acompanhamento_id)
        for row in rows:
            anexo_id = int(row["id"])
            self.acompanhamento_attachment_rows[anexo_id] = row
            self.acompanhamento_anexos_listbox.insert("end", f"{anexo_id} - {Path(row['arquivo_copiado']).name}")
        self._update_acompanhamento_summary()

    def refresh_socializacao_movimentacoes_list(self) -> None:
        self.socializacao_movimentacoes_rows = []
        self.socializacao_movimentacoes_listbox.delete(0, "end")
        if not self.selected_socializacao_id:
            self._update_socializacao_summary()
            return
        rows = self.database.list_socializacao_movimentacoes(self.selected_socializacao_id)
        self.socializacao_movimentacoes_rows = rows
        for row in rows:
            data_label = self._format_date_for_display(str(row["data_movimentacao"]), full_year=True)
            resumo = str(row["descricao"] or "").strip().replace("\n", " ")
            label = f"{data_label} - {resumo[:80]}"
            self.socializacao_movimentacoes_listbox.insert("end", label)
        self._update_socializacao_summary()

    def refresh_socializacao_attachments_list(self) -> None:
        self.socializacao_attachment_rows = {}
        self.socializacao_anexos_listbox.delete(0, "end")
        if not self.selected_socializacao_id:
            self._update_socializacao_summary()
            return
        rows = self.database.list_socializacao_anexos(self.selected_socializacao_id)
        for row in rows:
            anexo_id = int(row["id"])
            self.socializacao_attachment_rows[anexo_id] = row
            self.socializacao_anexos_listbox.insert("end", f"{anexo_id} - {Path(row['arquivo_copiado']).name}")
        self._update_socializacao_summary()

    def _update_acompanhamento_summary(self, row: sqlite3.Row | None = None) -> None:
        if not hasattr(self, "acompanhamento_summary_labels"):
            return
        if row is None and self.selected_acompanhamento_id:
            row = self.database.get_acompanhamento(self.selected_acompanhamento_id)
        if not row:
            for label in self.acompanhamento_summary_labels.values():
                label.configure(text="-")
            return
        abertura = self._format_date_for_display(str(row["data_abertura"]), full_year=True)
        resolucao = (
            self._format_date_for_display(str(row["data_resolucao"]), full_year=True)
            if row["data_resolucao"]
            else "-"
        )
        self.acompanhamento_summary_labels["status"].configure(
            text=f"{self._display_label(str(row['status']))} | {self._display_label(str(row['prioridade']))}"
        )
        self.acompanhamento_summary_labels["periodo"].configure(
            text=f"{abertura} -> {resolucao}"
        )
        self.acompanhamento_summary_labels["movimentacoes"].configure(
            text=str(len(self.acompanhamento_movimentacoes_rows))
        )
        self.acompanhamento_summary_labels["anexos"].configure(
            text=str(len(self.acompanhamento_attachment_rows))
        )

    def _update_socializacao_summary(self, row: dict[str, object] | sqlite3.Row | None = None) -> None:
        if not hasattr(self, "socializacao_summary_labels"):
            return
        if row is None and self.selected_socializacao_id:
            row = self.database.get_socializacao(self.selected_socializacao_id)
        if not row:
            for label in self.socializacao_summary_labels.values():
                label.configure(text="-")
            return
        if isinstance(row, sqlite3.Row):
            status = str(row["status_envio"] or "")
            data_envio = str(row["data_envio"] or "")
            apoio = bool(row["necessita_apoio"])
            destaque = bool(row["destaque_potencial"])
        else:
            status = str(row.get("status_envio", "") or "")
            data_envio = str(row.get("data_envio", "") or "")
            apoio = bool(row.get("necessita_apoio"))
            destaque = bool(row.get("destaque_potencial"))
        envio = self._format_date_for_display(data_envio, full_year=True) if data_envio else "-"
        self.socializacao_summary_labels["status"].configure(
            text=(
                f"{self._display_label(status)} | "
                f"apoio={'sim' if apoio else 'não'} | "
                f"destaque={'sim' if destaque else 'não'}"
            )
        )
        self.socializacao_summary_labels["envio"].configure(text=envio)
        self.socializacao_summary_labels["movimentacoes"].configure(
            text=str(len(self.socializacao_movimentacoes_rows))
        )
        self.socializacao_summary_labels["anexos"].configure(
            text=str(len(self.socializacao_attachment_rows))
        )

    def _get_selected_acompanhamento_movimentacao_row(self) -> sqlite3.Row | None:
        selection = self.acompanhamento_movimentacoes_listbox.curselection()
        if not selection:
            return None
        index = selection[0]
        if index >= len(self.acompanhamento_movimentacoes_rows):
            return None
        return self.acompanhamento_movimentacoes_rows[index]

    def _get_selected_socializacao_movimentacao_row(self) -> sqlite3.Row | None:
        selection = self.socializacao_movimentacoes_listbox.curselection()
        if not selection:
            return None
        index = selection[0]
        if index >= len(self.socializacao_movimentacoes_rows):
            return None
        return self.socializacao_movimentacoes_rows[index]

    def _get_selected_attachment_row(
        self,
        listbox: tk.Listbox,
        row_map: dict[int, sqlite3.Row],
    ) -> sqlite3.Row | None:
        selection = listbox.curselection()
        if not selection:
            return None
        item_text = listbox.get(selection[0])
        try:
            row_id = int(item_text.split(" - ", 1)[0])
        except ValueError:
            return None
        return row_map.get(row_id)

    def _ensure_acompanhamento_saved(self) -> int | None:
        if self.selected_acompanhamento_id:
            return self.selected_acompanhamento_id
        return self.save_acompanhamento(show_message=False)

    def _ensure_socializacao_saved(self) -> int | None:
        if self.selected_socializacao_id:
            return self.selected_socializacao_id
        return self.save_socializacao(show_message=False)

    def clear_cursista_form(self) -> None:
        self.selected_cursista_id = None
        self.cursista_nome_var.set("")
        self.cursista_turma_var.set("")
        self.cursista_email_institucional_var.set("")
        self.cursista_email_pessoal_var.set("")
        self.cursista_telefone_var.set("")
        self.cursista_status_busca_ativa_var.set(self._display_label(BUSCA_ATIVA_STATUS_OPTIONS[0]))
        self.cursista_ativo_var.set(True)
        self.cursista_observacao_text.delete("1.0", "end")

    def save_cursista(self) -> None:
        nome = self.cursista_nome_var.get().strip()
        if not nome:
            messagebox.showwarning("Cursista incompleto", "Informe o nome do cursista.")
            return

        turma_label = self.cursista_turma_var.get().strip()
        payload = {
            "nome": nome,
            "turma_id": self.turma_label_to_id.get(turma_label) if turma_label else None,
            "email_institucional": self.cursista_email_institucional_var.get().strip(),
            "email_pessoal": self.cursista_email_pessoal_var.get().strip(),
            "telefone_whatsapp": self.cursista_telefone_var.get().strip(),
            "status_busca_ativa": self._internal_value_from_label(
                self.cursista_status_busca_ativa_var.get().strip()
            )
            or BUSCA_ATIVA_STATUS_OPTIONS[0],
            "observacao_geral": self.cursista_observacao_text.get("1.0", "end").strip(),
            "ativo": self.cursista_ativo_var.get(),
        }
        self.selected_cursista_id = self.database.save_cursista(payload, self.selected_cursista_id)
        self._refresh_cursista_combo_options()
        self.refresh_cursista_tree(select_id=self.selected_cursista_id)
        self.refresh_acompanhamento_tree()
        self.refresh_socializacao_tree()
        self._refresh_cursistas_dashboard()
        self.status_bar_var.set("Cursista salvo no banco local.")
        messagebox.showinfo("Cursista salvo", "Cadastro de cursista salvo com sucesso.")

    def refresh_cursista_tree(self, select_id: int | None = None) -> None:
        include_inactive = self.cursista_filter_ativo_var.get() == "todos"
        turma_label = self.cursista_filter_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label) if turma_label else None
        rows = self.database.list_cursistas(
            include_inactive=include_inactive,
            turma_id=turma_id,
            search=self.cursista_filter_nome_var.get(),
            status_busca_ativa=self._internal_value_from_label(
                self.cursista_filter_busca_var.get().strip()
            ),
        )
        for item in self.cursista_tree.get_children():
            self.cursista_tree.delete(item)

        for row in rows:
            self.cursista_tree.insert(
                "",
                "end",
                iid=str(int(row["id"])),
                values=(
                    row["nome"],
                    row["turma_codigo"] or "",
                    self._display_label(str(row["status_busca_ativa"])),
                    self._mask_email(str(row["email_institucional"] or "")),
                    self._mask_phone(str(row["telefone_whatsapp"] or "")),
                    "sim" if int(row["ativo"] or 0) else "não",
                ),
            )

        if select_id is not None and str(select_id) in self.cursista_tree.get_children():
            self.cursista_tree.selection_set(str(select_id))
            self.cursista_tree.focus(str(select_id))

    def load_selected_cursista(self) -> None:
        selection = self.cursista_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione um cursista", "Escolha um cursista na lista.")
            return
        cursista_id = int(selection[0])
        row = self.database.get_cursista(cursista_id)
        if not row:
            return
        self.selected_cursista_id = cursista_id
        self.cursista_nome_var.set(row["nome"])
        self.cursista_turma_var.set(row["turma_codigo"] or "")
        self.cursista_email_institucional_var.set(row["email_institucional"])
        self.cursista_email_pessoal_var.set(row["email_pessoal"])
        self.cursista_telefone_var.set(row["telefone_whatsapp"])
        self.cursista_status_busca_ativa_var.set(self._display_label(str(row["status_busca_ativa"])))
        self.cursista_ativo_var.set(bool(row["ativo"]))
        self.cursista_observacao_text.delete("1.0", "end")
        self.cursista_observacao_text.insert("1.0", row["observacao_geral"])

    def deactivate_selected_cursista(self) -> None:
        selection = self.cursista_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione um cursista", "Escolha um cursista para inativar.")
            return
        cursista_id = int(selection[0])
        row = self.database.get_cursista(cursista_id)
        if not row:
            return
        if not messagebox.askyesno(
            "Inativar cursista",
            f"Deseja inativar o cursista {row['nome']}?",
        ):
            return
        self.database.deactivate_cursista(cursista_id)
        self.clear_cursista_form()
        self._refresh_cursista_combo_options()
        self.refresh_cursista_tree()
        self.refresh_acompanhamento_tree()
        self.refresh_socializacao_tree()
        self._refresh_cursistas_dashboard()
        self.status_bar_var.set("Cursista inativado.")

    def _load_cursista_import_rows(self, file_path: Path) -> list[dict[str, object]]:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError(
                "A leitura de planilhas Excel requer a dependência openpyxl."
            ) from error

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                return []

            normalized_headers = [self._normalize_import_header(header) for header in headers]
            header_aliases = {
                "nome": "nome",
                "email": "email",
                "e mail": "email",
                "telefone": "telefone",
                "telefone whatsapp": "telefone",
                "whatsapp": "telefone",
                "turma": "turma",
                "busca ativa": "status_busca_ativa",
                "status busca ativa": "status_busca_ativa",
                "observacao geral": "observacao_geral",
                "observacoes": "observacao_geral",
                "ativo": "ativo",
            }

            column_map: dict[str, int] = {}
            for index, header in enumerate(normalized_headers):
                key = header_aliases.get(header)
                if key and key not in column_map:
                    column_map[key] = index

            missing = [field for field in ("nome", "turma") if field not in column_map]
            if missing:
                missing_labels = ", ".join(missing)
                raise ValueError(
                    f"A planilha precisa conter as colunas obrigatórias: {missing_labels}."
                )

            imported_rows: list[dict[str, object]] = []
            for excel_row_number, row in enumerate(rows_iter, start=2):
                imported_rows.append(
                    {
                        "source_row": excel_row_number,
                        "nome": row[column_map["nome"]] if column_map.get("nome") is not None and column_map["nome"] < len(row) else "",
                        "email": row[column_map["email"]] if column_map.get("email") is not None and column_map["email"] < len(row) else "",
                        "telefone": row[column_map["telefone"]] if column_map.get("telefone") is not None and column_map["telefone"] < len(row) else "",
                        "turma": row[column_map["turma"]] if column_map.get("turma") is not None and column_map["turma"] < len(row) else "",
                        "status_busca_ativa": row[column_map["status_busca_ativa"]] if column_map.get("status_busca_ativa") is not None and column_map["status_busca_ativa"] < len(row) else "",
                        "observacao_geral": row[column_map["observacao_geral"]] if column_map.get("observacao_geral") is not None and column_map["observacao_geral"] < len(row) else "",
                        "ativo": row[column_map["ativo"]] if column_map.get("ativo") is not None and column_map["ativo"] < len(row) else "",
                    }
                )
            return imported_rows
        finally:
            workbook.close()

    def _show_cursista_import_preview_dialog(
        self,
        file_path: Path,
        imported_rows: list[dict[str, object]],
        preview: dict[str, object],
    ) -> None:
        dialog = tk.Toplevel(self)
        dialog.title(f"Prévia da planilha - {file_path.name}")
        dialog.geometry("1080x620")
        dialog.minsize(920, 520)
        dialog.transient(self)
        dialog.grab_set()

        container = ttk.Frame(dialog, padding=12)
        container.pack(fill="both", expand=True)

        header = ttk.LabelFrame(container, text="Resumo da análise", padding=10)
        header.pack(fill="x")

        totals = dict(preview.get("totals", {}))
        summary_rows = [
            ("Arquivo", file_path.name),
            ("Novos", str(totals.get("novo", 0))),
            ("Atualizações", str(totals.get("atualizar", 0))),
            ("Conflitos", str(totals.get("conflito", 0))),
            ("Ignorados", str(totals.get("ignorado", 0))),
            ("Processáveis", str(preview.get("processavel", 0))),
        ]
        for index, (label_text, value_text) in enumerate(summary_rows):
            ttk.Label(header, text=label_text).grid(row=index // 3, column=(index % 3) * 2, sticky="w", pady=2)
            ttk.Label(header, text=value_text).grid(
                row=index // 3,
                column=(index % 3) * 2 + 1,
                sticky="w",
                padx=(10, 24),
                pady=2,
            )

        info_var = tk.StringVar(
            value="Revise os dados antes de importar. Somente linhas processáveis serão gravadas."
        )
        ttk.Label(container, textvariable=info_var).pack(anchor="w", pady=(10, 8))

        tree_frame = ttk.LabelFrame(container, text="Linhas da planilha", padding=8)
        tree_frame.pack(fill="both", expand=True)
        columns = ("linha", "status", "motivo", "nome", "turma", "email", "telefone")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=18)
        headers = {
            "linha": "Linha",
            "status": "Status",
            "motivo": "Motivo",
            "nome": "Nome",
            "turma": "Turma",
            "email": "E-mail",
            "telefone": "Telefone",
        }
        min_widths = {
            "linha": 60,
            "status": 95,
            "motivo": 180,
            "nome": 240,
            "turma": 70,
            "email": 180,
            "telefone": 120,
        }
        weights = {
            "linha": 6,
            "status": 10,
            "motivo": 18,
            "nome": 28,
            "turma": 8,
            "email": 18,
            "telefone": 12,
        }
        for column in columns:
            tree.heading(column, text=headers[column])
            tree.column(column, width=min_widths[column], anchor="w", stretch=True)
        tree.pack(side="left", fill="both", expand=True)
        self._register_treeview_autofit(tree, columns, weights, min_widths)

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scroll_y.pack(side="right", fill="y")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        scroll_x.pack(side="bottom", fill="x")
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        tree.tag_configure("novo", background="#ecfdf5")
        tree.tag_configure("atualizar", background="#eff6ff")
        tree.tag_configure("conflito", background="#fef2f2")
        tree.tag_configure("ignorado", background="#f9fafb")

        for item in preview.get("rows", []):
            status = str(item.get("status") or "")
            tree.insert(
                "",
                "end",
                values=(
                    item.get("source_row", ""),
                    self._display_import_preview_status(status),
                    self._display_import_preview_reason(str(item.get("motivo") or "")),
                    str(item.get("nome") or ""),
                    str(item.get("turma_codigo") or ""),
                    str(item.get("email_pessoal") or ""),
                    str(item.get("telefone_whatsapp") or ""),
                ),
                tags=(status,),
            )

        def apply_import() -> None:
            processable = int(preview.get("processavel", 0) or 0)
            if processable <= 0:
                messagebox.showwarning(
                    "Nada para importar",
                    "Não há linhas processáveis nesta planilha.",
                    parent=dialog,
                )
                return

            totals_local = dict(preview.get("totals", {}))
            if not messagebox.askyesno(
                "Confirmar importação",
                (
                    f"Arquivo: {file_path.name}\n\n"
                    f"Novos cursistas: {totals_local.get('novo', 0)}\n"
                    f"Atualizações: {totals_local.get('atualizar', 0)}\n"
                    f"Conflitos ignorados: {totals_local.get('conflito', 0)}\n"
                    f"Ignorados: {totals_local.get('ignorado', 0)}\n\n"
                    "Deseja importar agora apenas as linhas processáveis?"
                ),
                parent=dialog,
            ):
                return

            result = self.database.apply_cursista_import(imported_rows)
            self.clear_cursista_form()
            self._refresh_cursista_combo_options()
            self.refresh_cursista_tree()
            self.refresh_acompanhamento_tree()
            self.refresh_socializacao_tree()
            self._refresh_cursistas_dashboard()
            self.status_bar_var.set(
                f"Importação concluída: {result.get('inseridos', 0)} inseridos, {result.get('atualizados', 0)} atualizados."
            )
            messagebox.showinfo(
                "Importação concluída",
                (
                    f"Arquivo: {file_path.name}\n\n"
                    f"Inseridos: {result.get('inseridos', 0)}\n"
                    f"Atualizados: {result.get('atualizados', 0)}\n"
                    f"Conflitos mantidos fora da importação: {totals_local.get('conflito', 0)}\n"
                    f"Ignorados: {totals_local.get('ignorado', 0)}"
                ),
                parent=dialog,
            )
            dialog.destroy()

        footer = ttk.Frame(container)
        footer.pack(fill="x", pady=(10, 0))
        import_button = ttk.Button(
            footer,
            text="Importar processáveis",
            command=apply_import,
        )
        import_button.pack(side="left")
        if int(preview.get("processavel", 0) or 0) <= 0:
            import_button.state(["disabled"])
        ttk.Button(footer, text="Fechar", command=dialog.destroy).pack(side="right")

        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)

    def preview_cursista_import_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha de cursistas",
            filetypes=[
                ("Planilhas Excel", "*.xlsx *.xlsm"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if not file_path:
            return

        selected_file = Path(file_path)
        try:
            imported_rows = self._load_cursista_import_rows(selected_file)
            preview = self.database.preview_cursista_import(imported_rows)
        except RuntimeError as error:
            messagebox.showerror("Dependência ausente", str(error))
            return
        except ValueError as error:
            messagebox.showwarning("Planilha inválida", str(error))
            return
        except Exception as error:
            messagebox.showerror("Falha na leitura", f"Não foi possível analisar a planilha.\n\n{error}")
            return

        self._show_cursista_import_preview_dialog(selected_file, imported_rows, preview)
        self.status_bar_var.set(
            f"Prévia da planilha concluída: {selected_file.name} ({preview.get('processavel', 0)} processáveis)."
        )

    def clear_acompanhamento_form(self) -> None:
        self.selected_acompanhamento_id = None
        self.acompanhamento_cursista_var.set("")
        self.acompanhamento_turma_var.set("")
        self.acompanhamento_encontro_var.set("")
        self.acompanhamento_categoria_var.set(self._display_label(ACOMPANHAMENTO_CATEGORIAS[0]))
        self.acompanhamento_status_var.set(self._display_label(ACOMPANHAMENTO_STATUS_OPTIONS[0]))
        self.acompanhamento_prioridade_var.set(self._display_label(PRIORIDADE_OPTIONS[1]))
        self.acompanhamento_resumo_var.set("")
        self.acompanhamento_data_abertura_var.set(datetime.today().strftime("%d/%m/%Y"))
        self.acompanhamento_data_resolucao_var.set("")
        self.acompanhamento_encaminhar_pec_var.set(False)
        self.acompanhamento_descricao_text.delete("1.0", "end")
        self.acompanhamento_movimentacao_data_var.set(datetime.today().strftime("%d/%m/%Y"))
        self.acompanhamento_movimentacao_text.delete("1.0", "end")
        self._populate_acompanhamento_encontros_for_turma(None)
        self.refresh_acompanhamento_movimentacoes_list()
        self.refresh_acompanhamento_attachments_list()

    def _on_acompanhamento_cursista_selected(self, _event=None) -> None:
        label = self.acompanhamento_cursista_var.get().strip()
        cursista_id = self.cursista_label_to_id.get(label)
        if not cursista_id:
            return
        row = self.database.get_cursista(cursista_id)
        if not row:
            return
        turma_codigo = str(row["turma_codigo"] or "").strip()
        self.acompanhamento_turma_var.set(turma_codigo)
        turma_id = self.turma_label_to_id.get(turma_codigo) if turma_codigo else None
        self._populate_acompanhamento_encontros_for_turma(turma_id)

    def _on_acompanhamento_turma_selected(self, _event=None) -> None:
        turma_label = self.acompanhamento_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label) if turma_label else None
        self._refresh_cursista_combo_options()
        self._populate_acompanhamento_encontros_for_turma(turma_id)

    def _on_socializacao_turma_selected(self, _event=None) -> None:
        self._refresh_cursista_combo_options()

    def save_acompanhamento(self, show_message: bool = True) -> int | None:
        cursista_label = self.acompanhamento_cursista_var.get().strip()
        cursista_id = self.cursista_label_to_id.get(cursista_label)
        if not cursista_id:
            messagebox.showwarning("Ocorrência incompleta", "Selecione o cursista.")
            return None
        resumo = self.acompanhamento_resumo_var.get().strip()
        if not resumo:
            messagebox.showwarning("Ocorrência incompleta", "Informe um resumo da ocorrência.")
            return None

        turma_label = self.acompanhamento_turma_var.get().strip()
        categoria = self._internal_value_from_label(
            self.acompanhamento_categoria_var.get().strip()
        ) or ACOMPANHAMENTO_CATEGORIAS[0]
        status = self._internal_value_from_label(
            self.acompanhamento_status_var.get().strip()
        ) or ACOMPANHAMENTO_STATUS_OPTIONS[0]
        prioridade = self._internal_value_from_label(
            self.acompanhamento_prioridade_var.get().strip()
        ) or PRIORIDADE_OPTIONS[1]
        data_resolucao = self.acompanhamento_data_resolucao_var.get().strip()
        if status == "resolvido" and not data_resolucao:
            data_resolucao = datetime.today().strftime("%d/%m/%Y")
            self.acompanhamento_data_resolucao_var.set(data_resolucao)

        try:
            payload = {
                "cursista_id": cursista_id,
                "turma_id": self.turma_label_to_id.get(turma_label) if turma_label else None,
                "encontro_id": self.acompanhamento_encontro_label_to_id.get(
                    self.acompanhamento_encontro_var.get().strip()
                ),
                "categoria": categoria,
                "status": status,
                "prioridade": prioridade,
                "resumo": resumo,
                "descricao": self.acompanhamento_descricao_text.get("1.0", "end").strip(),
                "encaminhar_pec": self.acompanhamento_encaminhar_pec_var.get(),
                "data_abertura": self._parse_date(self.acompanhamento_data_abertura_var.get()),
                "data_resolucao": self._parse_optional_date(data_resolucao),
            }
        except ValueError as error:
            messagebox.showwarning("Data inválida", str(error))
            return None

        self.selected_acompanhamento_id = self.database.save_acompanhamento(
            payload,
            self.selected_acompanhamento_id,
        )
        self.refresh_acompanhamento_tree(select_id=self.selected_acompanhamento_id)
        self.refresh_acompanhamento_movimentacoes_list()
        self.refresh_acompanhamento_attachments_list()
        self._refresh_cursistas_dashboard()
        self.status_bar_var.set("Ocorrência do cursista salva.")
        if show_message:
            messagebox.showinfo("Ocorrência salva", "Registro salvo com sucesso.")
        return self.selected_acompanhamento_id

    def refresh_acompanhamento_tree(self, select_id: int | None = None) -> None:
        self._refresh_cursista_combo_options()
        turma_label = self.acompanhamento_filter_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label) if turma_label else None
        rows = self.database.list_acompanhamentos(
            status_filter=self._internal_value_from_label(
                self.acompanhamento_filter_status_var.get().strip()
            ),
            turma_id=turma_id,
            categoria=self._internal_value_from_label(
                self.acompanhamento_filter_categoria_var.get().strip()
            ),
            search=self.acompanhamento_filter_search_var.get().strip(),
            attachment_filter=self._internal_value_from_label(
                self.acompanhamento_filter_anexos_var.get().strip()
            ),
            movement_filter=self._internal_value_from_label(
                self.acompanhamento_filter_movimentacoes_var.get().strip()
            ),
        )
        for item in self.acompanhamento_tree.get_children():
            self.acompanhamento_tree.delete(item)

        for row in rows:
            tags: tuple[str, ...] = ()
            if row["prioridade"] == "alta" and row["status"] not in {"resolvido", "arquivado"}:
                tags = ("high_priority",)
            elif row["status"] in {"resolvido", "arquivado"}:
                tags = ("resolved",)
            else:
                tags = ("pending",)
            self.acompanhamento_tree.insert(
                "",
                "end",
                iid=str(int(row["id"])),
                values=(
                    self._format_date_for_display(str(row["data_abertura"]), full_year=True),
                    row["cursista_nome"],
                    row["turma_codigo"] or "",
                    self._display_label(str(row["categoria"])),
                    self._display_label(str(row["status"])),
                    self._display_label(str(row["prioridade"])),
                ),
                tags=tags,
            )

        if select_id is not None and str(select_id) in self.acompanhamento_tree.get_children():
            self.acompanhamento_tree.selection_set(str(select_id))
            self.acompanhamento_tree.focus(str(select_id))
        elif not self.acompanhamento_tree.get_children():
            self._update_acompanhamento_summary(None)

    def load_selected_acompanhamento(self) -> None:
        selection = self.acompanhamento_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione uma ocorrência", "Escolha uma ocorrência na lista.")
            return
        acompanhamento_id = int(selection[0])
        row = self.database.get_acompanhamento(acompanhamento_id)
        if not row:
            return
        self.selected_acompanhamento_id = acompanhamento_id
        cursista_row = self.database.get_cursista(int(row["cursista_id"]))
        if cursista_row:
            self.acompanhamento_cursista_var.set(self._format_cursista_label(cursista_row))
        self.acompanhamento_turma_var.set(row["turma_codigo"] or "")
        turma_id = int(row["turma_id"]) if row["turma_id"] else None
        encontro_row = self.database.get_encontro(int(row["encontro_id"])) if row["encontro_id"] else None
        month_id = int(encontro_row["month_id"]) if encontro_row else self.current_month_id
        self._populate_acompanhamento_encontros_for_turma(turma_id, month_id)
        if encontro_row:
            self.acompanhamento_encontro_var.set(self._format_encontro_label(encontro_row))
        else:
            self.acompanhamento_encontro_var.set("")
        self.acompanhamento_categoria_var.set(self._display_label(str(row["categoria"] or "")))
        self.acompanhamento_status_var.set(self._display_label(str(row["status"] or "")))
        self.acompanhamento_prioridade_var.set(self._display_label(str(row["prioridade"] or "")))
        self.acompanhamento_resumo_var.set(row["resumo"])
        self.acompanhamento_data_abertura_var.set(
            self._format_date_for_display(str(row["data_abertura"]), full_year=True)
        )
        self.acompanhamento_data_resolucao_var.set(
            self._format_date_for_display(str(row["data_resolucao"]), full_year=True)
            if row["data_resolucao"]
            else ""
        )
        self.acompanhamento_encaminhar_pec_var.set(bool(row["encaminhar_pec"]))
        self.acompanhamento_descricao_text.delete("1.0", "end")
        self.acompanhamento_descricao_text.insert("1.0", row["descricao"])
        self.refresh_acompanhamento_movimentacoes_list()
        self.refresh_acompanhamento_attachments_list()
        self._update_acompanhamento_summary(row)

    def add_acompanhamento_movimentacao(self) -> None:
        acompanhamento_id = self._ensure_acompanhamento_saved()
        if not acompanhamento_id:
            return
        descricao = self.acompanhamento_movimentacao_text.get("1.0", "end").strip()
        if not descricao:
            messagebox.showwarning("Movimentação vazia", "Informe a movimentação a registrar.")
            return
        try:
            data_movimentacao = self._parse_date(self.acompanhamento_movimentacao_data_var.get())
        except ValueError as error:
            messagebox.showwarning("Data inválida", str(error))
            return
        self.database.add_acompanhamento_movimentacao(acompanhamento_id, data_movimentacao, descricao)
        self.acompanhamento_movimentacao_text.delete("1.0", "end")
        self.acompanhamento_movimentacao_data_var.set(datetime.today().strftime("%d/%m/%Y"))
        self.refresh_acompanhamento_movimentacoes_list()
        self.status_bar_var.set("Movimentação adicionada à ocorrência.")

    def remove_selected_acompanhamento_movimentacao(self) -> None:
        row = self._get_selected_acompanhamento_movimentacao_row()
        if not row:
            messagebox.showwarning("Selecione uma movimentação", "Escolha uma movimentação para remover.")
            return
        if not messagebox.askyesno("Remover movimentação", "Deseja remover a movimentação selecionada?"):
            return
        self.database.delete_acompanhamento_movimentacao(int(row["id"]))
        self.refresh_acompanhamento_movimentacoes_list()
        self.status_bar_var.set("Movimentação removida.")

    def add_acompanhamento_attachment(self) -> None:
        acompanhamento_id = self._ensure_acompanhamento_saved()
        if not acompanhamento_id:
            return
        file_paths = filedialog.askopenfilenames(title="Selecione arquivos da ocorrência")
        if not file_paths:
            return
        folder = self._build_acompanhamento_attachment_directory(acompanhamento_id)
        saved = 0
        for source in file_paths:
            source_path = Path(source)
            target = self._next_attachment_path(folder, source_path.name)
            shutil.copy2(source_path, target)
            self.database.add_acompanhamento_anexo(acompanhamento_id, str(source_path), str(target))
            saved += 1
        self.refresh_acompanhamento_attachments_list()
        self.status_bar_var.set(f"{saved} arquivo(s) anexado(s) à ocorrência.")

    def open_selected_acompanhamento_attachment(self) -> None:
        row = self._get_selected_attachment_row(
            self.acompanhamento_anexos_listbox,
            self.acompanhamento_attachment_rows,
        )
        if not row:
            messagebox.showwarning("Selecione um arquivo", "Escolha um anexo para abrir.")
            return
        self._open_path(Path(row["arquivo_copiado"]))

    def remove_selected_acompanhamento_attachment(self) -> None:
        row = self._get_selected_attachment_row(
            self.acompanhamento_anexos_listbox,
            self.acompanhamento_attachment_rows,
        )
        if not row:
            messagebox.showwarning("Selecione um arquivo", "Escolha um anexo para remover.")
            return
        if not messagebox.askyesno("Remover arquivo", "Deseja remover o anexo selecionado?"):
            return
        removed_path = self.database.delete_acompanhamento_anexo(int(row["id"]))
        if removed_path:
            self._delete_file_if_exists(removed_path)
        self.refresh_acompanhamento_attachments_list()
        self.status_bar_var.set("Anexo da ocorrência removido.")

    def _get_socializacao_reference(self, show_message: bool = True) -> tuple[int, int] | None:
        try:
            month = int(self.socializacao_mes_var.get().strip())
            year = int(self.socializacao_ano_var.get().strip())
        except ValueError:
            if show_message:
                messagebox.showwarning(
                    "Período inválido",
                    "Informe mês e ano válidos para a socialização.",
                )
            return None
        if month < 1 or month > 12:
            if show_message:
                messagebox.showwarning("Mês inválido", "Informe um mês entre 1 e 12.")
            return None
        return month, year

    def clear_socializacao_form(self) -> None:
        self.selected_socializacao_id = None
        self.socializacao_cursista_var.set("")
        self.socializacao_turma_var.set("")
        self.socializacao_status_var.set(self._display_label(SOCIALIZACAO_STATUS_OPTIONS[0]))
        self.socializacao_data_envio_var.set("")
        self.socializacao_necessita_apoio_var.set(False)
        self.socializacao_destaque_var.set(False)
        self.socializacao_observacao_text.delete("1.0", "end")
        self.socializacao_movimentacao_data_var.set(datetime.today().strftime("%d/%m/%Y"))
        self.socializacao_movimentacao_text.delete("1.0", "end")
        self.refresh_socializacao_movimentacoes_list()
        self.refresh_socializacao_attachments_list()

    def _on_socializacao_cursista_selected(self, _event=None) -> None:
        label = self.socializacao_cursista_var.get().strip()
        cursista_id = self.cursista_label_to_id.get(label)
        if not cursista_id:
            return
        row = self.database.get_cursista(cursista_id)
        if not row:
            return
        self.socializacao_turma_var.set(row["turma_codigo"] or "")

    def save_socializacao(self, show_message: bool = True) -> int | None:
        cursista_label = self.socializacao_cursista_var.get().strip()
        cursista_id = self.cursista_label_to_id.get(cursista_label)
        if not cursista_id:
            messagebox.showwarning("Socialização incompleta", "Selecione o cursista.")
            return None
        reference = self._get_socializacao_reference()
        if not reference:
            return None
        month, year = reference
        turma_label = self.socializacao_turma_var.get().strip()

        try:
            payload = {
                "cursista_id": cursista_id,
                "turma_id": self.turma_label_to_id.get(turma_label) if turma_label else None,
                "mes_referencia": month,
                "ano_referencia": year,
                "status_envio": self._internal_value_from_label(
                    self.socializacao_status_var.get().strip()
                )
                or SOCIALIZACAO_STATUS_OPTIONS[0],
                "data_envio": self._parse_optional_date(self.socializacao_data_envio_var.get()),
                "observacao_pedagogica": self.socializacao_observacao_text.get("1.0", "end").strip(),
                "necessita_apoio": self.socializacao_necessita_apoio_var.get(),
                "destaque_potencial": self.socializacao_destaque_var.get(),
            }
        except ValueError as error:
            messagebox.showwarning("Data inválida", str(error))
            return None

        self.selected_socializacao_id = self.database.upsert_socializacao_mensal(payload)
        self.refresh_socializacao_tree(select_item=f"s{self.selected_socializacao_id}")
        self.refresh_socializacao_movimentacoes_list()
        self.refresh_socializacao_attachments_list()
        self._refresh_cursistas_dashboard()
        self.status_bar_var.set("Socialização registrada.")
        if show_message:
            messagebox.showinfo("Socialização salva", "Registro salvo com sucesso.")
        return self.selected_socializacao_id

    def refresh_socializacao_tree(
        self,
        select_item: str | None = None,
        show_message: bool = False,
    ) -> None:
        self._refresh_cursista_combo_options()
        reference = self._get_socializacao_reference(show_message=show_message)
        if not reference:
            return
        month, year = reference
        turma_label = self.socializacao_filter_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label) if turma_label else None
        rows = self.database.list_socializacoes(
            month,
            year,
            turma_id=turma_id,
            status_envio=self._internal_value_from_label(
                self.socializacao_filter_status_var.get().strip()
            ),
            search=self.socializacao_filter_search_var.get().strip(),
            attachment_filter=self._internal_value_from_label(
                self.socializacao_filter_anexos_var.get().strip()
            ),
            movement_filter=self._internal_value_from_label(
                self.socializacao_filter_movimentacoes_var.get().strip()
            ),
        )
        self.socializacao_rows_by_item_id = {}
        for item in self.socializacao_tree.get_children():
            self.socializacao_tree.delete(item)

        for row in rows:
            item_id = f"s{row['socializacao_id']}" if row["socializacao_id"] else f"c{row['cursista_id']}"
            self.socializacao_rows_by_item_id[item_id] = dict(row)
            tags: tuple[str, ...] = ()
            if int(row["destaque_potencial"] or 0):
                tags = ("highlight",)
            elif int(row["necessita_apoio"] or 0):
                tags = ("support",)
            elif str(row["status_envio"]) == "nao_enviada":
                tags = ("pending",)
            self.socializacao_tree.insert(
                "",
                "end",
                iid=item_id,
                values=(
                    row["cursista_nome"],
                    row["turma_codigo"] or "",
                    self._display_label(str(row["status_envio"])),
                    self._format_date_for_display(str(row["data_envio"]), full_year=True)
                    if row["data_envio"]
                    else "",
                    "sim" if int(row["necessita_apoio"] or 0) else "não",
                    "sim" if int(row["destaque_potencial"] or 0) else "não",
                ),
                tags=tags,
            )

        self._refresh_cursistas_dashboard()
        if select_item is not None and select_item in self.socializacao_tree.get_children():
            self.socializacao_tree.selection_set(select_item)
            self.socializacao_tree.focus(select_item)
        elif not self.socializacao_tree.get_children():
            self._update_socializacao_summary(None)

    def load_selected_socializacao(self) -> None:
        selection = self.socializacao_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione uma socialização", "Escolha um registro na lista.")
            return
        item_id = selection[0]
        row = self.socializacao_rows_by_item_id.get(item_id)
        if not row:
            return
        self.selected_socializacao_id = int(row["socializacao_id"]) if row["socializacao_id"] else None
        cursista_row = self.database.get_cursista(int(row["cursista_id"]))
        if cursista_row:
            self.socializacao_cursista_var.set(self._format_cursista_label(cursista_row))
        self.socializacao_turma_var.set(str(row["turma_codigo"] or ""))
        self.socializacao_status_var.set(
            self._display_label(str(row["status_envio"] or SOCIALIZACAO_STATUS_OPTIONS[0]))
        )
        self.socializacao_data_envio_var.set(
            self._format_date_for_display(str(row["data_envio"]), full_year=True)
            if row["data_envio"]
            else ""
        )
        self.socializacao_necessita_apoio_var.set(bool(row["necessita_apoio"]))
        self.socializacao_destaque_var.set(bool(row["destaque_potencial"]))
        self.socializacao_observacao_text.delete("1.0", "end")
        self.socializacao_observacao_text.insert("1.0", str(row["observacao_pedagogica"] or ""))
        self.refresh_socializacao_movimentacoes_list()
        self.refresh_socializacao_attachments_list()
        self._update_socializacao_summary(row)

    def mark_socializacao_sent_today(self) -> None:
        self.socializacao_status_var.set(self._display_label("enviada"))
        self.socializacao_data_envio_var.set(datetime.today().strftime("%d/%m/%Y"))

    def add_socializacao_movimentacao(self) -> None:
        socializacao_id = self._ensure_socializacao_saved()
        if not socializacao_id:
            return
        descricao = self.socializacao_movimentacao_text.get("1.0", "end").strip()
        if not descricao:
            messagebox.showwarning("Movimentação vazia", "Informe a movimentação a registrar.")
            return
        try:
            data_movimentacao = self._parse_date(self.socializacao_movimentacao_data_var.get())
        except ValueError as error:
            messagebox.showwarning("Data inválida", str(error))
            return
        self.database.add_socializacao_movimentacao(socializacao_id, data_movimentacao, descricao)
        self.socializacao_movimentacao_text.delete("1.0", "end")
        self.socializacao_movimentacao_data_var.set(datetime.today().strftime("%d/%m/%Y"))
        self.refresh_socializacao_movimentacoes_list()
        self.status_bar_var.set("Movimentação adicionada à socialização.")

    def remove_selected_socializacao_movimentacao(self) -> None:
        row = self._get_selected_socializacao_movimentacao_row()
        if not row:
            messagebox.showwarning("Selecione uma movimentação", "Escolha uma movimentação para remover.")
            return
        if not messagebox.askyesno("Remover movimentação", "Deseja remover a movimentação selecionada?"):
            return
        self.database.delete_socializacao_movimentacao(int(row["id"]))
        self.refresh_socializacao_movimentacoes_list()
        self.status_bar_var.set("Movimentação removida.")

    def add_socializacao_attachment(self) -> None:
        socializacao_id = self._ensure_socializacao_saved()
        if not socializacao_id:
            return
        file_paths = filedialog.askopenfilenames(title="Selecione arquivos da socialização")
        if not file_paths:
            return
        folder = self._build_socializacao_attachment_directory(socializacao_id)
        saved = 0
        for source in file_paths:
            source_path = Path(source)
            target = self._next_attachment_path(folder, source_path.name)
            shutil.copy2(source_path, target)
            self.database.add_socializacao_anexo(socializacao_id, str(source_path), str(target))
            saved += 1
        self.refresh_socializacao_attachments_list()
        self.status_bar_var.set(f"{saved} arquivo(s) anexado(s) à socialização.")

    def open_selected_socializacao_attachment(self) -> None:
        row = self._get_selected_attachment_row(
            self.socializacao_anexos_listbox,
            self.socializacao_attachment_rows,
        )
        if not row:
            messagebox.showwarning("Selecione um arquivo", "Escolha um anexo para abrir.")
            return
        self._open_path(Path(row["arquivo_copiado"]))

    def remove_selected_socializacao_attachment(self) -> None:
        row = self._get_selected_attachment_row(
            self.socializacao_anexos_listbox,
            self.socializacao_attachment_rows,
        )
        if not row:
            messagebox.showwarning("Selecione um arquivo", "Escolha um anexo para remover.")
            return
        if not messagebox.askyesno("Remover arquivo", "Deseja remover o anexo selecionado?"):
            return
        removed_path = self.database.delete_socializacao_anexo(int(row["id"]))
        if removed_path:
            self._delete_file_if_exists(removed_path)
        self.refresh_socializacao_attachments_list()
        self.status_bar_var.set("Anexo da socialização removido.")

    def collect_professor_data(self) -> dict[str, str]:
        return {
            "nome": self.nome_var.get().strip(),
            "tema": self.tema_var.get().strip(),
            "email_institucional": self.email_var.get().strip(),
            "diretoria_ure": self.diretoria_var.get().strip(),
            "pec_responsavel": self.pec_var.get().strip(),
            "valor_formacao_semanal": self.valor_formacao_semanal_var.get().strip(),
        }

    def save_current_month(self) -> None:
        if not self.current_month_id:
            messagebox.showwarning("Mês não selecionado", "Crie ou selecione um mês primeiro.")
            return
        if not self._normalize_currency_field():
            return
        rate_decision = self._confirm_month_rate_update(self.current_month_id)
        if rate_decision is None:
            return
        self.database.save_month(
            self.current_month_id,
            self.collect_professor_data(),
            update_weekly_rate=rate_decision,
        )
        self.status_bar_var.set("Dados do mês atualizados.")
        messagebox.showinfo("Dados salvos", "Os dados do professor foram salvos no mês ativo.")

    def save_defaults(self) -> None:
        if not self._normalize_currency_field():
            return
        self.database.save_defaults(self.collect_professor_data())
        self.status_bar_var.set("Padrão do professor atualizado.")
        messagebox.showinfo("Padrão salvo", "Os dados foram gravados como padrão para próximos meses.")

    def apply_defaults_to_form(self) -> None:
        defaults = self.database.get_defaults()
        self.nome_var.set(defaults.get("nome", ""))
        self.tema_var.set(defaults.get("tema", ""))
        self.email_var.set(defaults.get("email_institucional", ""))
        self.diretoria_var.set(defaults.get("diretoria_ure", ""))
        self.pec_var.set(defaults.get("pec_responsavel", ""))
        self.valor_formacao_semanal_var.set(defaults.get("valor_formacao_semanal", ""))

    def clear_turma_form(self) -> None:
        self.selected_turma_id = None
        self.turma_codigo_var.set("")
        self.turma_dia_var.set(WEEKDAY_OPTIONS[0])
        self.turma_horario_var.set("")
        self.turma_componente_var.set("")
        self.turma_situacao_var.set(TURMA_STATUS_OPTIONS[0])

    def save_turma(self) -> None:
        codigo = self.turma_codigo_var.get().strip()
        if not codigo:
            messagebox.showwarning("Turma incompleta", "Informe o código da turma.")
            return

        payload = {
            "codigo": codigo,
            "dia_semana": self.turma_dia_var.get().strip(),
            "horario": self.turma_horario_var.get().strip(),
            "componente": self.turma_componente_var.get().strip(),
            "situacao": self.turma_situacao_var.get().strip() or TURMA_STATUS_OPTIONS[0],
        }

        try:
            self.selected_turma_id = self.database.save_turma(payload, self.selected_turma_id)
        except sqlite3.IntegrityError:
            messagebox.showerror("Turma duplicada", "Já existe uma turma com esse código.")
            return

        self.refresh_turma_tree(select_id=self.selected_turma_id)
        self.refresh_encontro_tree()
        self._refresh_cursista_combo_options()
        self.refresh_cursista_tree()
        self.refresh_acompanhamento_tree()
        self.refresh_socializacao_tree()
        self.status_bar_var.set("Turma atualizada no banco local.")
        messagebox.showinfo("Turma salva", "Cadastro de turma salvo com sucesso.")

    def refresh_turma_tree(self, select_id: int | None = None) -> None:
        rows = self.database.list_turmas(include_inactive=True)
        for item in self.turma_tree.get_children():
            self.turma_tree.delete(item)

        active_labels = []
        self.turma_label_to_id = {}
        for row in rows:
            turma_id = int(row["id"])
            values = (
                row["codigo"],
                row["dia_semana"],
                row["horario"],
                row["componente"],
                row["situacao"],
            )
            self.turma_tree.insert("", "end", iid=str(turma_id), values=values)
            if row["situacao"] == "ativa":
                active_labels.append(row["codigo"])
                self.turma_label_to_id[row["codigo"]] = turma_id
            elif row["codigo"] not in self.turma_label_to_id:
                self.turma_label_to_id[row["codigo"]] = turma_id

        self.encontro_turma_combo["values"] = active_labels or list(self.turma_label_to_id.keys())
        self._refresh_cursista_module_turma_options()

        if select_id is not None and str(select_id) in self.turma_tree.get_children():
            self.turma_tree.selection_set(str(select_id))
            self.turma_tree.focus(str(select_id))

    def load_selected_turma(self) -> None:
        selection = self.turma_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione uma turma", "Escolha uma turma na lista.")
            return

        turma_id = int(selection[0])
        row = self.database.get_turma(turma_id)
        if not row:
            return

        self.selected_turma_id = turma_id
        self.turma_codigo_var.set(row["codigo"])
        self.turma_dia_var.set(row["dia_semana"])
        self.turma_horario_var.set(row["horario"])
        self.turma_componente_var.set(row["componente"])
        self.turma_situacao_var.set(row["situacao"])

    def toggle_selected_turma_status(self) -> None:
        selection = self.turma_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione uma turma", "Escolha uma turma para alterar a situação.")
            return

        turma_id = int(selection[0])
        row = self.database.get_turma(turma_id)
        if not row:
            return

        new_status = "inativa" if row["situacao"] == "ativa" else "ativa"
        payload = {
            "codigo": row["codigo"],
            "dia_semana": row["dia_semana"],
            "horario": row["horario"],
            "componente": row["componente"],
            "situacao": new_status,
        }
        self.database.save_turma(payload, turma_id)
        self.refresh_turma_tree(select_id=turma_id)
        self._refresh_cursista_combo_options()
        self.refresh_cursista_tree()
        self.refresh_acompanhamento_tree()
        self.refresh_socializacao_tree()
        self.status_bar_var.set(f"Situação da turma alterada para {new_status}.")

    def clear_encontro_form(self) -> None:
        self.selected_encontro_id = None
        self.selected_evidence_id = None
        self.encontro_data_var.set("")
        self.encontro_pauta_var.set("")
        self.encontro_turma_var.set("")
        self.encontro_inicio_var.set("")
        self.encontro_fim_var.set("")
        self.encontro_participantes_var.set("")
        self.encontro_duracao_var.set("")
        self.encontro_status_var.set(STATUS_OPTIONS[0])
        self.observacao_text.delete("1.0", "end")
        self.evidence_listbox.delete(0, "end")
        self._clear_preview("Selecione uma evidência à esquerda.")

    def apply_auto_observation(self) -> None:
        selected = self.auto_observacao_var.get()
        text = AUTO_OBSERVATIONS.get(selected)
        if not text:
            return
        self.observacao_text.delete("1.0", "end")
        self.observacao_text.insert("1.0", text)

    def _parse_date(self, raw_value: str) -> str:
        text = raw_value.strip()
        if not text:
            raise ValueError("Não permitir encontro sem data.")

        for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, date_format).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError("Use a data no formato dd/mm/aaaa ou aaaa-mm-dd.")

    def _normalize_time(self, raw_value: str) -> str:
        text = raw_value.strip()
        if not text:
            return ""
        try:
            return datetime.strptime(text, "%H:%M").strftime("%H:%M")
        except ValueError as error:
            raise ValueError("Os horários precisam usar o formato HH:MM.") from error

    def _get_or_create_month_from_iso_date(self, iso_date: str) -> int:
        encounter_date = datetime.strptime(iso_date, "%Y-%m-%d")
        existing_month = self.database.get_month_by_reference(encounter_date.month, encounter_date.year)
        month_id = self.database.get_or_create_month(encounter_date.month, encounter_date.year)
        if not self._normalize_currency_field():
            raise ValueError("Informe um valor válido para a formação semanal.")
        update_weekly_rate = True
        if existing_month:
            rate_decision = self._confirm_month_rate_update(month_id)
            if rate_decision is None:
                raise ValueError("Operação cancelada pelo usuário.")
            update_weekly_rate = rate_decision
        self.database.save_month(
            month_id,
            self.collect_professor_data(),
            update_weekly_rate=update_weekly_rate,
        )
        return month_id

    def _collect_encontro_payload(self) -> dict[str, object]:
        turma_label = self.encontro_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label)
        if not turma_id:
            raise ValueError("Não permitir encontro sem turma.")

        pauta_text = self.encontro_pauta_var.get().strip()
        if not pauta_text:
            raise ValueError("Não permitir encontro sem número de pauta.")

        try:
            pauta_numero = int(pauta_text)
        except ValueError as error:
            raise ValueError("O número da pauta precisa ser numérico.") from error

        participantes_text = self.encontro_participantes_var.get().strip()
        participantes = None
        if participantes_text:
            try:
                participantes = int(participantes_text)
            except ValueError as error:
                raise ValueError("Participantes precisa ser um numero inteiro.") from error

        data_encontro = self._parse_date(self.encontro_data_var.get())
        month_id = self._get_or_create_month_from_iso_date(data_encontro)

        return {
            "month_id": month_id,
            "turma_id": turma_id,
            "data_encontro": data_encontro,
            "pauta_numero": pauta_numero,
            "hora_inicio": self._normalize_time(self.encontro_inicio_var.get()),
            "hora_fim": self._normalize_time(self.encontro_fim_var.get()),
            "participantes": participantes,
            "duracao": self.encontro_duracao_var.get().strip(),
            "observacao": self.observacao_text.get("1.0", "end").strip(),
            "situacao": self.encontro_status_var.get().strip() or STATUS_OPTIONS[0],
        }

    def save_encontro(self, show_message: bool = True) -> int | None:
        try:
            payload = self._collect_encontro_payload()
        except ValueError as error:
            messagebox.showwarning("Encontro incompleto", str(error))
            return None

        try:
            self.selected_encontro_id = self.database.save_encontro(payload, self.selected_encontro_id)
        except sqlite3.IntegrityError:
            messagebox.showerror(
                "Encontro duplicado",
                "Já existe encontro com a mesma data, turma e pauta neste mês.",
            )
            return None

        target_month_id = int(payload["month_id"])
        self.refresh_month_selector(select_month_id=target_month_id)
        self.load_month(target_month_id)
        encontro_id = self.selected_encontro_id
        if encontro_id:
            self.load_encontro_into_form(encontro_id)

        if show_message:
            messagebox.showinfo("Encontro salvo", "Registro do encontro salvo com sucesso.")
        self.status_bar_var.set("Encontro salvo no banco local e vinculado ao mês da data.")
        return encontro_id

    def refresh_encontro_tree(self, select_id: int | None = None) -> None:
        for item in self.encontro_tree.get_children():
            self.encontro_tree.delete(item)

        if not self.current_month_id:
            return

        rows = self.database.list_encontros(self.current_month_id)
        for row in rows:
            self.encontro_tree.insert(
                "",
                "end",
                iid=str(int(row["id"])),
                values=(
                    self._format_date_for_display(row["data_encontro"]),
                    row["pauta_numero"],
                    row["turma_codigo"],
                    row["situacao"],
                    row["participantes"] if row["participantes"] is not None else "",
                    row["total_imagens"],
                ),
            )

        if select_id is not None and str(select_id) in self.encontro_tree.get_children():
            self.encontro_tree.selection_set(str(select_id))
            self.encontro_tree.focus(str(select_id))

    def load_selected_encontro(self) -> None:
        selection = self.encontro_tree.selection()
        if not selection:
            messagebox.showwarning("Selecione um encontro", "Escolha um encontro na lista.")
            return
        self.load_encontro_into_form(int(selection[0]))

    def load_encontro_into_form(self, encontro_id: int) -> None:
        row = self.database.get_encontro(encontro_id)
        if not row:
            return

        self.selected_encontro_id = encontro_id
        self.encontro_data_var.set(self._format_date_for_display(row["data_encontro"], full_year=True))
        self.encontro_pauta_var.set(str(row["pauta_numero"]))
        self.encontro_turma_var.set(row["turma_codigo"])
        self.encontro_inicio_var.set(row["hora_inicio"])
        self.encontro_fim_var.set(row["hora_fim"])
        self.encontro_participantes_var.set(
            "" if row["participantes"] is None else str(row["participantes"])
        )
        self.encontro_duracao_var.set(row["duracao"])
        self.encontro_status_var.set(row["situacao"])
        self.observacao_text.delete("1.0", "end")
        self.observacao_text.insert("1.0", row["observacao"])
        self.refresh_evidence_list()

    def _format_date_for_display(self, iso_date: str, full_year: bool = False) -> str:
        fmt = "%d/%m/%Y" if full_year else "%d/%m"
        return datetime.strptime(iso_date, "%Y-%m-%d").strftime(fmt)

    def _calculate_end_time(self, start_time: str) -> str:
        normalized_start = self._normalize_time(start_time)
        start_dt = datetime.strptime(normalized_start, "%H:%M")
        end_dt = start_dt + timedelta(minutes=90)
        return end_dt.strftime("%H:%M")

    def _fill_schedule_from_selected_turma(self) -> None:
        turma_label = self.encontro_turma_var.get().strip()
        turma_id = self.turma_label_to_id.get(turma_label)
        if not turma_id:
            return

        turma = self.database.get_turma(turma_id)
        if not turma:
            return

        horario_inicio = (turma["horario"] or "").strip()
        if not horario_inicio:
            self.encontro_inicio_var.set("")
            self.encontro_fim_var.set("")
            self.encontro_duracao_var.set("")
            return

        try:
            horario_inicio = self._normalize_time(horario_inicio)
            horario_fim = self._calculate_end_time(horario_inicio)
        except ValueError:
            self.status_bar_var.set(
                "A turma selecionada tem horário inválido. Ajuste o cadastro da turma."
            )
            return

        self.encontro_inicio_var.set(horario_inicio)
        self.encontro_fim_var.set(horario_fim)
        self.encontro_duracao_var.set("1h30")

    def _on_encontro_turma_selected(self, _event=None) -> None:
        self._fill_schedule_from_selected_turma()

    def ensure_encontro_saved(self) -> int | None:
        if self.selected_encontro_id:
            return self.selected_encontro_id
        return self.save_encontro(show_message=False)

    def _build_evidence_directory(self, encontro_row: sqlite3.Row) -> Path:
        month = self.database.get_month(int(encontro_row["month_id"]))
        if not month:
            raise ValueError("Mês do encontro não encontrado.")

        month_folder = f"{month['ref_year']}-{int(month['ref_month']):02d}"
        turma_folder = slugify(str(encontro_row["turma_codigo"]))
        pauta_folder = f"pauta_{int(encontro_row['pauta_numero']):02d}"
        target = EVIDENCIAS_DIR / month_folder / turma_folder / pauta_folder
        target.mkdir(parents=True, exist_ok=True)
        return target

    def _next_evidence_path(self, encontro_row: sqlite3.Row, extension: str) -> Path:
        folder = self._build_evidence_directory(encontro_row)
        data = str(encontro_row["data_encontro"])
        counter = 1
        while True:
            candidate = folder / f"{data}_{counter:02d}{extension}"
            if not candidate.exists():
                return candidate
            counter += 1

    def refresh_evidence_list(self) -> None:
        self.selected_evidence_id = None
        self.evidence_listbox.delete(0, "end")
        if not self.selected_encontro_id:
            self._clear_preview("Selecione uma evidência à esquerda.")
            return

        evidences = self.database.list_evidences(self.selected_encontro_id)
        for item in evidences:
            label = f"{item['id']} - {Path(item['arquivo_copiado']).name}"
            self.evidence_listbox.insert("end", label)

        if evidences:
            self.evidence_listbox.selection_set(0)
            self._on_evidence_selected()
        else:
            self._clear_preview("Nenhuma evidência anexada para este encontro.")

    def _clear_preview(self, message: str) -> None:
        self.preview_photo = None
        self.preview_image_path = None
        self.preview_title_var.set(message)
        self.preview_canvas.delete("all")

    def _on_evidence_selected(self, _event=None) -> None:
        evidence = self._get_selected_evidence_row()
        if not evidence:
            self._clear_preview("Selecione uma evidência à esquerda.")
            return

        image_path = Path(evidence["arquivo_copiado"])
        self.preview_image_path = image_path
        self.preview_title_var.set(image_path.name)
        self._render_preview_image()

    def _render_preview_image(self) -> None:
        image_path = getattr(self, "preview_image_path", None)
        if not image_path:
            return

        canvas_width = max(self.preview_canvas.winfo_width(), 320)
        canvas_height = max(self.preview_canvas.winfo_height(), 220)
        self.preview_canvas.delete("all")

        try:
            from PIL import Image, ImageTk
        except ImportError:
            self.preview_canvas.create_text(
                canvas_width / 2,
                canvas_height / 2,
                text="Instale Pillow para ver a pré-visualização.\n\npip install Pillow",
                justify="center",
            )
            return

        try:
            image = Image.open(image_path)
        except Exception as error:
            self.preview_canvas.create_text(
                canvas_width / 2,
                canvas_height / 2,
                text=f"Não foi possível abrir a imagem.\n\n{error}",
                justify="center",
            )
            return

        image.thumbnail((canvas_width - 16, canvas_height - 16))
        self.preview_photo = ImageTk.PhotoImage(image)
        self.preview_canvas.create_image(
            canvas_width / 2,
            canvas_height / 2,
            image=self.preview_photo,
            anchor="center",
        )

    def _get_selected_evidence_row(self) -> sqlite3.Row | None:
        if not self.selected_encontro_id:
            return None
        selection = self.evidence_listbox.curselection()
        if not selection:
            return None
        index = selection[0]
        evidences = self.database.list_evidences(self.selected_encontro_id)
        if index >= len(evidences):
            return None
        return evidences[index]

    def add_image_file(self) -> None:
        encontro_id = self.ensure_encontro_saved()
        if not encontro_id:
            return

        file_paths = filedialog.askopenfilenames(
            title="Selecione imagens",
            filetypes=[("Imagens", "*.png *.jpg *.jpeg *.bmp *.gif *.webp")],
        )
        if not file_paths:
            return

        encontro_row = self.database.get_encontro(encontro_id)
        if not encontro_row:
            return

        saved = 0
        for source in file_paths:
            source_path = Path(source)
            extension = source_path.suffix.lower() or ".png"
            target = self._next_evidence_path(encontro_row, extension)
            shutil.copy2(source_path, target)
            self.database.add_evidence(encontro_id, str(source_path), str(target))
            saved += 1

        self.refresh_evidence_list()
        self.refresh_encontro_tree(select_id=encontro_id)
        self.refresh_checklist()
        self.status_bar_var.set(f"{saved} imagem(ns) anexada(s) ao encontro.")

    def paste_image_from_clipboard(self) -> None:
        encontro_id = self.ensure_encontro_saved()
        if not encontro_id:
            return

        encontro_row = self.database.get_encontro(encontro_id)
        if not encontro_row:
            return

        try:
            from PIL import ImageGrab
        except ImportError:
            messagebox.showerror(
                "Dependência ausente",
                "O recurso de colar print precisa da biblioteca Pillow.\n\n"
                "Instale com:\n"
                "pip install Pillow",
            )
            return

        try:
            clipboard_data = ImageGrab.grabclipboard()
        except Exception as error:
            messagebox.showerror("Área de transferência", f"Não foi possível acessar o clipboard: {error}")
            return

        if clipboard_data is None:
            messagebox.showwarning("Clipboard vazio", "Não há imagem copiada no momento.")
            return

        saved = 0
        if isinstance(clipboard_data, list):
            for item in clipboard_data:
                path = Path(item)
                if path.suffix.lower() not in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"}:
                    continue
                target = self._next_evidence_path(encontro_row, path.suffix.lower())
                shutil.copy2(path, target)
                self.database.add_evidence(encontro_id, str(path), str(target))
                saved += 1
        else:
            target = self._next_evidence_path(encontro_row, ".png")
            clipboard_data.save(target, format="PNG")
            self.database.add_evidence(encontro_id, "clipboard", str(target))
            saved = 1

        if saved == 0:
            messagebox.showwarning(
                "Clipboard sem imagem válida",
                "O conteúdo copiado não contém imagem ou arquivo de imagem suportado.",
            )
            return

        self.refresh_evidence_list()
        self.refresh_encontro_tree(select_id=encontro_id)
        self.refresh_checklist()
        self.status_bar_var.set("Imagem colada e salva localmente.")

    def open_selected_evidence(self) -> None:
        row = self._get_selected_evidence_row()
        if not row:
            messagebox.showwarning("Selecione uma imagem", "Escolha uma evidência na lista.")
            return
        self._open_path(Path(row["arquivo_copiado"]))

    def remove_selected_evidence(self) -> None:
        row = self._get_selected_evidence_row()
        if not row:
            messagebox.showwarning("Selecione uma imagem", "Escolha uma evidência para remover.")
            return

        if not messagebox.askyesno("Remover imagem", "Deseja remover a evidência selecionada"):
            return

        removed_path = self.database.delete_evidence(int(row["id"]))
        if removed_path:
            try:
                Path(removed_path).unlink(missing_ok=True)
            except TypeError:
                path = Path(removed_path)
                if path.exists():
                    path.unlink()

        self.refresh_evidence_list()
        if self.selected_encontro_id:
            self.refresh_encontro_tree(select_id=self.selected_encontro_id)
        self.refresh_checklist()
        self.status_bar_var.set("Evidência removida.")

    def refresh_checklist(self) -> None:
        checklist = self.database.get_checklist(self.current_month_id) if self.current_month_id else None
        if not checklist:
            for label in self.checklist_labels.values():
                label.configure(text="-")
            self.checklist_listbox.delete(0, "end")
            return

        self.checklist_labels["total_registrados"].configure(text=str(checklist["total_registrados"]))
        self.checklist_labels["total_realizados"].configure(text=str(checklist["total_realizados"]))
        self.checklist_labels["valor_formacao_semanal"].configure(
            text=self._format_currency_from_cents(int(checklist["valor_formacao_semanal_centavos"]))
        )
        self.checklist_labels["valor_total_mensal"].configure(
            text=self._format_currency_from_cents(int(checklist["valor_total_mensal_centavos"]))
        )
        self.checklist_labels["turmas"].configure(
            text=", ".join(checklist["turmas"]) if checklist["turmas"] else "-"
        )
        self.checklist_labels["pautas"].configure(
            text=", ".join(str(item) for item in checklist["pautas"]) if checklist["pautas"] else "-"
        )
        self.checklist_labels["sem_imagem"].configure(text=str(len(checklist["sem_imagem"])))
        self.checklist_labels["sem_observacao"].configure(text=str(len(checklist["sem_observacao"])))
        self.checklist_labels["sem_horario"].configure(text=str(len(checklist["sem_horario"])))
        self.checklist_labels["sem_participantes"].configure(
            text=str(len(checklist["sem_participantes"]))
        )

        self.checklist_listbox.delete(0, "end")
        issues = checklist["issues"] or ["Nenhuma pendência crítica encontrada."]
        for issue in issues:
            self.checklist_listbox.insert("end", issue)

        detail_groups = [
            ("Por turma", checklist["por_turma"]),
            ("Por pauta", checklist["por_pauta"]),
            ("Por situação", checklist["por_status"]),
        ]
        for title, values in detail_groups:
            self.checklist_listbox.insert("end", "")
            self.checklist_listbox.insert("end", title)
            for key, value in values.items():
                self.checklist_listbox.insert("end", f"  {key}: {value}")

        missing_groups = [
            ("Encontros sem imagem", checklist["sem_imagem"]),
            ("Encontros sem observação", checklist["sem_observacao"]),
            ("Encontros sem horário", checklist["sem_horario"]),
            ("Encontros sem participantes", checklist["sem_participantes"]),
        ]
        for title, items in missing_groups:
            if not items:
                continue
            self.checklist_listbox.insert("end", "")
            self.checklist_listbox.insert("end", title)
            for item in items:
                self.checklist_listbox.insert("end", f"  {item}")

    def _generate_bundle(self) -> dict[str, object] | None:
        if not self.current_month_id:
            messagebox.showwarning("Mês não selecionado", "Crie ou selecione um mês primeiro.")
            return None

        if not self._normalize_currency_field():
            return None
        self.database.save_month(self.current_month_id, self.collect_professor_data())
        return self.database.get_month_bundle(self.current_month_id)

    def _set_report_generation_state(self, in_progress: bool, message: str = "") -> None:
        self.report_generation_in_progress = in_progress
        button_state = "disabled" if in_progress else "normal"
        self.generate_docx_button.configure(state=button_state)
        self.generate_pdf_button.configure(state=button_state)
        self.generate_financial_button.configure(state=button_state)
        self.refresh_checklist_button.configure(state=button_state)
        self.open_evidence_button.configure(state=button_state)
        self.open_output_button.configure(state=button_state)

        if in_progress:
            self.report_progress_var.set(message)
            self.report_progress.pack(anchor="w", pady=(10, 0))
            self.report_progress.start(12)
        else:
            self.report_progress.stop()
            self.report_progress.pack_forget()
            self.report_progress_var.set(message)

    def _start_report_generation(self, report_type: str) -> None:
        if self.report_generation_in_progress:
            messagebox.showinfo(
                "Processando relatório",
                "Já existe uma geração de relatório em andamento. Aguarde a conclusão.",
            )
            return

        bundle = self._generate_bundle()
        if not bundle:
            return

        progress_messages = {
            "docx": "Gerando relatório DOCX...",
            "pdf": "Gerando relatório PDF...",
            "financial": "Gerando extrato financeiro...",
        }
        progress_message = progress_messages.get(report_type, "Processando arquivo...")
        self._set_report_generation_state(True, progress_message)
        self.status_bar_var.set(progress_message)

        worker = threading.Thread(
            target=self._run_report_generation_task,
            args=(report_type, bundle),
            daemon=True,
        )
        worker.start()

    def _run_report_generation_task(self, report_type: str, bundle: dict[str, object]) -> None:
        try:
            if report_type == "financial":
                output_path = generate_financial_statement_docx(
                    month_data=bundle["month"],
                    encontros=bundle["encontros"],
                    output_dir=SAIDAS_DIR,
                )
            else:
                docx_path = generate_docx(
                    month_data=bundle["month"],
                    encontros=bundle["encontros"],
                    turmas=bundle["turmas"],
                    total_realizados=bundle["checklist"]["total_realizados"],
                    output_dir=SAIDAS_DIR,
                )
                output_path = docx_path
            if report_type == "pdf":
                output_path = export_pdf(docx_path)
        except Exception as error:
            self.after(0, lambda err=error, kind=report_type: self._finish_report_generation_error(kind, err))
            return

        self.after(0, lambda path=output_path, kind=report_type: self._finish_report_generation_success(kind, path))

    def _finish_report_generation_success(self, report_type: str, output_path: Path) -> None:
        self._set_report_generation_state(False, "")
        if report_type == "pdf":
            self.status_bar_var.set(f"Relatório PDF gerado em {output_path.name}.")
            messagebox.showinfo(
                "PDF pronto",
                "Relatório gerado com sucesso.\n\n"
                f"Arquivo:\n{output_path}\n\n"
                "Você pode encontrá-lo na pasta 'saídas' do programa.\n"
                "Se preferir, clique no botão 'Abrir saídas' para acessar a pasta.",
            )
        elif report_type == "financial":
            self.status_bar_var.set(f"Extrato financeiro gerado em {output_path.name}.")
            messagebox.showinfo(
                "Extrato financeiro pronto",
                "Extrato gerado com sucesso.\n\n"
                f"Arquivo:\n{output_path}\n\n"
                "Você pode encontrá-lo na pasta 'saídas' do programa.\n"
                "Se preferir, clique no botão 'Abrir saídas' para acessar a pasta.",
            )
        else:
            self.status_bar_var.set(f"Relatório DOCX gerado em {output_path.name}.")
            messagebox.showinfo(
                "Relatório pronto",
                "Relatório gerado com sucesso.\n\n"
                f"Arquivo:\n{output_path}\n\n"
                "Você pode encontrá-lo na pasta 'saídas' do programa.\n"
                "Se preferir, clique no botão 'Abrir saídas' para acessar a pasta.",
            )

    def _finish_report_generation_error(self, report_type: str, error: Exception) -> None:
        self._set_report_generation_state(False, "")
        if report_type == "pdf":
            messagebox.showerror(
                "PDF indisponível",
                "Não foi possível exportar para PDF.\n\n"
                f"Detalhe: {error}\n\n"
                "Instale o LibreOffice ou use o Microsoft Word com pywin32.",
            )
            self.status_bar_var.set("Falha ao gerar relatório PDF.")
        elif report_type == "financial":
            messagebox.showerror("Falha ao gerar extrato", str(error))
            self.status_bar_var.set("Falha ao gerar extrato financeiro.")
        else:
            messagebox.showerror("Falha ao gerar DOCX", str(error))
            self.status_bar_var.set("Falha ao gerar relatório DOCX.")

    def generate_docx_report(self) -> None:
        self._start_report_generation("docx")

    def generate_pdf_report(self) -> None:
        self._start_report_generation("pdf")

    def generate_financial_report(self) -> None:
        self._start_report_generation("financial")

    def show_about_dialog(self) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Sobre o projeto")
        dialog.geometry("760x560")
        dialog.minsize(640, 440)
        dialog.resizable(False, False)
        dialog.transient(self)

        container = ttk.Frame(dialog, padding=16)
        container.pack(fill="both", expand=True)

        ttk.Label(
            container,
            text="Multiplica Evidências",
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w")
        ttk.Label(
            container,
            text=f"Versão {APP_VERSION}",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(2, 2))
        ttk.Label(
            container,
            text="Júlio César Valera • Professor de Matemática, Programação e Robótica",
            font=("Segoe UI", 10, "italic"),
        ).pack(anchor="w", pady=(4, 14))
        ttk.Label(
            container,
            text=(
                "Contato: julio@projetos.tec.br • "
                "juliovalera@professor.educacao.sp.gov.br"
            ),
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(0, 10))
        ttk.Label(
            container,
            text=self._get_terms_acceptance_summary(),
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(0, 10))
        tk.Label(
            container,
            text=(
                "AVISO IMPORTANTE: PROJETO INDEPENDENTE E NÃO OFICIAL.\n"
                "Uso experimental possível em encontros pedagógicos, inclusive em contextos como o\n"
                "Programa Multiplica, sem homologação ou vínculo institucional automático."
            ),
            font=("Segoe UI", 10, "bold"),
            fg="#8b1e1e",
            bg=self._ttk_background(),
            justify="left",
        ).pack(anchor="w", pady=(0, 10))

        text_frame = ttk.Frame(container)
        text_frame.pack(fill="both", expand=True)

        scrollbar = ttk.Scrollbar(text_frame, orient="vertical")
        scrollbar.pack(side="right", fill="y")

        text_widget = tk.Text(
            text_frame,
            wrap="word",
            font=("Segoe UI", 10),
            padx=12,
            pady=12,
            yscrollcommand=scrollbar.set,
            relief="solid",
            borderwidth=1,
            background="#fcfcfd",
        )
        text_widget.pack(fill="both", expand=True)
        scrollbar.config(command=text_widget.yview)

        text_widget.insert("1.0", self._get_about_text())
        text_widget.config(state="disabled")

        button_row = ttk.Frame(container)
        button_row.pack(fill="x", pady=(12, 0))
        ttk.Button(
            button_row,
            text="Copiar e-mails",
            command=self.copy_about_contacts,
        ).pack(side="left")
        ttk.Button(
            button_row,
            text="Ajuda",
            command=self.show_help_dialog,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(
            button_row,
            text="Ler termos novamente",
            command=self.open_terms_review_dialog,
        ).pack(side="left", padx=(10, 0))
        ttk.Button(button_row, text="Fechar", command=dialog.destroy).pack(side="right")

        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        dialog.focus_set()

    def copy_about_contacts(self) -> None:
        contacts = "julio@projetos.tec.br; juliovalera@professor.educacao.sp.gov.br"
        self.clipboard_clear()
        self.clipboard_append(contacts)
        self.update_idletasks()
        messagebox.showinfo(
            "Contatos copiados",
            "Os e-mails de contato foram copiados para a área de transferência.",
        )

    def show_help_dialog(self) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Ajuda")
        dialog.geometry("820x620")
        dialog.minsize(700, 520)
        dialog.transient(self)

        container = ttk.Frame(dialog, padding=16)
        container.pack(fill="both", expand=True)

        ttk.Label(
            container,
            text="Ajuda e manual do usuário",
            font=("Segoe UI", 15, "bold"),
        ).pack(anchor="w")
        ttk.Label(
            container,
            text="Consulte o guia rápido, o manual completo e a versão editorial do manual.",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(4, 12))

        actions = ttk.Frame(container)
        actions.pack(fill="x", pady=(0, 12))
        ttk.Button(actions, text="Abrir guia rápido", command=self.open_quick_guide).pack(side="left")
        ttk.Button(actions, text="Abrir manual completo", command=self.open_user_manual).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(actions, text="Abrir manual editorial", command=self.open_editorial_manual).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(actions, text="Abrir pasta docs", command=self.open_docs_folder).pack(
            side="left", padx=(8, 0)
        )

        text_frame = ttk.Frame(container)
        text_frame.pack(fill="both", expand=True)

        scrollbar = ttk.Scrollbar(text_frame, orient="vertical")
        scrollbar.pack(side="right", fill="y")

        text_widget = tk.Text(
            text_frame,
            wrap="word",
            font=("Segoe UI", 10),
            padx=12,
            pady=12,
            yscrollcommand=scrollbar.set,
            relief="solid",
            borderwidth=1,
            background="#fcfcfd",
        )
        text_widget.pack(fill="both", expand=True)
        scrollbar.config(command=text_widget.yview)
        text_widget.insert("1.0", self._load_help_text())
        text_widget.config(state="disabled")

        footer = ttk.Frame(container)
        footer.pack(fill="x", pady=(12, 0))
        ttk.Button(footer, text="Fechar", command=dialog.destroy).pack(side="right")

        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        dialog.focus_set()

    def open_quick_guide(self) -> None:
        self._open_document(GUIA_RAPIDO_PATH, "Guia rápido")

    def open_user_manual(self) -> None:
        self._open_document(MANUAL_USUARIO_PATH, "Manual do usuário")

    def open_editorial_manual(self) -> None:
        self._open_document(MANUAL_EDITORIAL_PATH, "Manual editorial")

    def open_docs_folder(self) -> None:
        self._open_path(GUIA_RAPIDO_PATH.parent)

    def _load_help_text(self) -> str:
        if AJUDA_RESUMIDA_PATH.exists():
            try:
                return AJUDA_RESUMIDA_PATH.read_text(encoding="utf-8")
            except OSError:
                pass
        return (
            "Ajuda indisponível no momento.\n\n"
            "Abra a pasta docs para consultar os arquivos de apoio."
        )

    def _open_document(self, path: Path, document_name: str) -> None:
        if not path.exists():
            messagebox.showwarning(
                "Arquivo não encontrado",
                f"O arquivo de {document_name.lower()} não foi localizado em:\n{path}",
            )
            return
        self._open_path(path)

    def ensure_terms_acceptance(self) -> None:
        accepted_version = self.database.get_app_config_value(self.TERMS_CONFIG_KEY)
        if accepted_version == self.TERMS_VERSION:
            return
        self.show_terms_dialog(require_acceptance=True)

    def open_terms_review_dialog(self) -> None:
        self.show_terms_dialog(require_acceptance=False)

    def show_terms_dialog(self, require_acceptance: bool = True) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Termos de uso")
        dialog.geometry("820x620")
        dialog.minsize(720, 520)
        dialog.transient(self)
        if require_acceptance:
            dialog.grab_set()

        container = ttk.Frame(dialog, padding=16)
        container.pack(fill="both", expand=True)

        ttk.Label(
            container,
            text="Termos de uso e aviso importante",
            font=("Segoe UI", 15, "bold"),
        ).pack(anchor="w")
        ttk.Label(
            container,
            text="Leia com atenção antes de utilizar o sistema.",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(6, 12))
        tk.Label(
            container,
            text=(
                "AVISO IMPORTANTE: ESTE SOFTWARE É INDEPENDENTE E NÃO OFICIAL.\n"
                "Pode ser utilizado experimentalmente em encontros pedagógicos; qualquer referência ao Programa Multiplica é apenas exemplificativa e descritiva."
            ),
            font=("Segoe UI", 10, "bold"),
            fg="#8b1e1e",
            bg=self._ttk_background(),
            justify="left",
        ).pack(anchor="w", pady=(0, 12))

        text_frame = ttk.Frame(container)
        text_frame.pack(fill="both", expand=True)

        text_scroll = ttk.Scrollbar(text_frame, orient="vertical")
        text_scroll.pack(side="right", fill="y")

        text_widget = tk.Text(
            text_frame,
            wrap="word",
            font=("Segoe UI", 10),
            padx=12,
            pady=12,
            yscrollcommand=text_scroll.set,
            relief="solid",
            borderwidth=1,
            background="#fcfcfd",
        )
        text_widget.pack(fill="both", expand=True)
        text_scroll.config(command=text_widget.yview)
        text_widget.insert("1.0", self._get_terms_text())
        text_widget.config(state="disabled")

        button_row = ttk.Frame(container)
        button_row.pack(fill="x", pady=(14, 0))

        if require_acceptance:
            self.terms_accept_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(
                container,
                text=(
                    "Li e compreendi os termos de uso. Reconheço que devo conferir os dados "
                    "antes de qualquer uso oficial."
                ),
                variable=self.terms_accept_var,
                command=self._update_terms_accept_button_state,
            ).pack(anchor="w", pady=(12, 0))

            ttk.Label(
                container,
                text=(
                    "Sem esse aceite o programa será fechado. O aceite fica salvo apenas neste computador."
                ),
                justify="left",
            ).pack(anchor="w", pady=(8, 0))

            ttk.Button(button_row, text="Sair", command=lambda: self._decline_terms(dialog)).pack(
                side="left"
            )
            self.accept_terms_button = ttk.Button(
                button_row,
                text="Aceito os termos",
                command=lambda: self._accept_terms(dialog),
                state="disabled",
            )
            self.accept_terms_button.pack(side="right")
            dialog.protocol("WM_DELETE_WINDOW", lambda: self._decline_terms(dialog))
            dialog.bind("<Escape>", lambda _event: self._decline_terms(dialog))
        else:
            ttk.Label(
                button_row,
                text=self._get_terms_acceptance_summary(),
                justify="left",
            ).pack(side="left")
            ttk.Button(button_row, text="Fechar", command=dialog.destroy).pack(side="right")
            dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)
            dialog.bind("<Escape>", lambda _event: dialog.destroy())

        dialog.focus_force()

    def _update_terms_accept_button_state(self) -> None:
        if hasattr(self, "accept_terms_button"):
            state = "normal" if self.terms_accept_var.get() else "disabled"
            self.accept_terms_button.configure(state=state)

    def _accept_terms(self, dialog: tk.Toplevel) -> None:
        accepted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.database.set_app_config_value(self.TERMS_CONFIG_KEY, self.TERMS_VERSION)
        self.database.set_app_config_value(self.TERMS_ACCEPTED_AT_KEY, accepted_at)
        self.status_bar_var.set("Termos de uso aceitos neste computador.")
        dialog.destroy()

    def _decline_terms(self, dialog: tk.Toplevel) -> None:
        dialog.destroy()
        self.after(50, self.destroy)

    def _get_terms_acceptance_summary(self) -> str:
        accepted_at = self.database.get_app_config_value(self.TERMS_ACCEPTED_AT_KEY)
        if not accepted_at:
            return "Aceite dos termos: ainda não registrado neste computador."
        try:
            accepted_dt = datetime.strptime(accepted_at, "%Y-%m-%d %H:%M:%S")
            accepted_label = accepted_dt.strftime("%d/%m/%Y às %H:%M")
        except ValueError:
            accepted_label = accepted_at
        return f"Aceite dos termos registrado neste computador em: {accepted_label}"

    def _get_terms_text(self) -> str:
        return (
            "Este projeto tem finalidade pedagógica, organizacional e de apoio técnico, tendo sido "
            "desenvolvido para auxiliar o registro local de encontros, evidências e relatórios em "
            "diferentes contextos pedagógicos, inclusive no trabalho do professor multiplicador.\n\n"
            "Este software pode ser utilizado experimentalmente em encontros pedagógicos diversos, inclusive em "
            "contextos como o Programa Multiplica. Qualquer menção a esse nome tem caráter "
            "exclusivamente descritivo e exemplificativo, sem representar oficialidade, homologação, "
            "licenciamento de marca ou vínculo institucional automático.\n\n"
            "O sistema não substitui a conferência humana nem a validação institucional das informações, "
            "documentos, datas, imagens, horários, participantes e relatórios gerados. Antes de qualquer "
            "envio, protocolo ou uso oficial, o usuário deve revisar cuidadosamente todo o conteúdo.\n\n"
            "Os dados permanecem armazenados localmente neste computador, incluindo banco SQLite, imagens "
            "e arquivos gerados, sem envio automático para serviços externos pelo sistema.\n\n"
            "Por se tratar de um projeto em desenvolvimento contínuo, o autor não garante adequação integral "
            "a normas, rotinas, exigências administrativas específicas ou funcionamento livre de falhas em "
            "todos os ambientes de uso. O sistema deve ser utilizado com prudência, bom senso e conferência "
            "final, especialmente em contextos institucionais.\n\n"
            "O aceite destes termos registra ciência e concordância com essas condições de uso. Em situações "
            "específicas, recomenda-se também observar as normas institucionais aplicáveis e, quando necessário, "
            "a orientação jurídica pertinente.\n\n"
            "Sugestões, críticas construtivas e contribuições são bem-vindas.\n\n"
            "Contato para diálogo e colaboração:\n"
            "• julio@projetos.tec.br\n"
            "• juliovalera@professor.educacao.sp.gov.br"
        )

    def _get_about_text(self) -> str:
        return (
            "Este projeto foi concebido com finalidade pedagógica, organizacional e de apoio técnico, "
            "buscando auxiliar o registro local de encontros, evidências e relatórios em diferentes "
            "contextos de trabalho pedagógico, inclusive no cotidiano do professor multiplicador. "
            "Pode ser utilizado experimentalmente, por exemplo, em contextos como o Programa Multiplica, sem que isso "
            "represente caráter oficial do software.\n\n"
            "Sua motivação nasce de uma necessidade concreta da rotina docente: organizar informações com mais "
            "clareza, menos retrabalho e maior autonomia no uso dos próprios dados.\n\n"
            "Ao reunir essas informações em um único ambiente local, o sistema procura valorizar o tempo "
            "pedagógico do professor multiplicador. Em vez de dispersar registros em pastas, anotações e "
            "documentos manuais, a proposta é oferecer um fluxo mais consistente, preservando a memória do "
            "trabalho realizado e facilitando a conferência mensal das evidências.\n\n"
            "A linguagem Python foi escolhida por reunir legibilidade, estabilidade, facilidade de manutenção "
            "e ótima integração com documentos, imagens e banco de dados local. Além disso, Python ocupa hoje "
            "um papel relevante na educação tecnológica, no ensino de programação e no desenvolvimento de "
            "materiais digitais conectados a problemas reais do cotidiano.\n\n"
            "Vivemos uma realidade em que materiais digitais, registros eletrônicos e produção documental fazem "
            "parte da vida escolar e profissional. Saber organizar dados com responsabilidade, preservar "
            "evidências e transformar informações em relatórios úteis é uma competência cada vez mais relevante. "
            "Este projeto se insere justamente nesse contexto: usar a tecnologia como instrumento de organização, "
            "autoria e apoio ao trabalho humano.\n\n"
            "Multiplica Evidências permanece em desenvolvimento contínuo. Por isso, o sistema não substitui a "
            "conferência humana nem a validação institucional das informações, documentos e relatórios gerados. "
            "Antes de qualquer envio, protocolo ou uso oficial, recomenda-se revisão cuidadosa de todo o conteúdo.\n\n"
            "Por se tratar de um projeto em evolução, não há garantia de adequação integral a todas as rotinas, "
            "normas ou exigências administrativas específicas, nem de funcionamento livre de falhas em todos os "
            "ambientes. O uso deve ocorrer com prudência, bom senso e conferência final, especialmente em contextos "
            "institucionais.\n\n"
            "Sugestões, críticas construtivas, observações de melhoria e contribuições de uso são bem-vindas. "
            "Este projeto também se fortalece pelo diálogo com outros educadores e usuários interessados em "
            "aperfeiçoar ferramentas digitais voltadas à organização pedagógica.\n\n"
            "Contato para diálogo e colaboração:\n"
            "• julio@projetos.tec.br\n"
            "• juliovalera@professor.educacao.sp.gov.br"
        )

    def open_output_folder(self) -> None:
        self._open_path(SAIDAS_DIR)

    def open_evidence_folder(self) -> None:
        self._open_path(EVIDENCIAS_DIR)

    def backup_database(self) -> None:
        backup_path = self.database.backup_database(BACKUP_DIR)
        self.status_bar_var.set(f"Backup criado: {backup_path.name}")
        messagebox.showinfo("Backup concluído", f"Arquivo salvo em:\n{backup_path}")

    def _open_path(self, path: Path) -> None:
        try:
            if os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.run(["open", str(path)], check=False)
            else:
                subprocess.run(["xdg-open", str(path)], check=False)
        except Exception as error:
            messagebox.showerror("Abrir arquivo", f"Não foi possível abrir:\n{path}\n\n{error}")
